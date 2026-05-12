import io
import json
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Avg, Count, DecimalField, ExpressionWrapper, F, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone

from accounting.models import CashRegister, Transaction as AccountingTransaction
from hr.models import DailyReport, Employee
from production.models import FinishedGoodsInventory, Product, ProductCategory
from .models import ReturnLog, Sale, SaleItem, ShiftClosure, ShiftDailyAllocation


def _can_access(user, *roles):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=roles).exists()


REGISTER_BY_PAYMENT = {
    "cash": "Main Cash",
    "terminal": "Terminal",
    "electronic": "Electronic",
}


def _decimal_zero():
    return Decimal("0.00")


def _sold_qty_shift_product(day, shift_name, product_id):
    agg = SaleItem.objects.filter(
        sale__date__date=day,
        sale__shift_name=shift_name,
        product_id=product_id,
    ).aggregate(t=Sum('quantity'))
    return int(agg['t'] or 0)


def _sale_total_units(sale):
    return int(
        SaleItem.objects.filter(sale_id=sale.id).aggregate(t=Sum('quantity'))['t'] or 0
    )


def _credit_piecework_for_sale(sale, employee_id=None):
    """Sotuv — HR ishbay avtomatik yozuv.

    Shartlar:
    - Agar formadan aniq `employee_id` kelsa, shu xodimga yoziladi.
    - Aks holda `sale.shift_name` bo'yicha qidiriladi.
    - Topilmasa `sale.seller_id` orqali `user_account` dan qidiriladi.
    """
    emp = None
    if employee_id:
        emp = Employee.objects.filter(id=employee_id, status='active', is_piecework=True).first()
    
    if not emp and sale.shift_name:
        emp = Employee.objects.filter(shift__name=sale.shift_name, status='active', is_piecework=True).first()

    if not emp and sale.seller_id:
        emp = Employee.objects.filter(user_account_id=sale.seller_id, status='active', is_piecework=True).first()

    if not emp:
        return
        
    sale_date = timezone.localtime(sale.date).date()
    total_qty = _sale_total_units(sale)
    if total_qty <= 0:
        return
        
    with transaction.atomic():
        report, created = DailyReport.objects.select_for_update().get_or_create(
            employee=emp,
            date=sale_date,
            shift=emp.shift,
            defaults={
                'units_produced': 0,
                'units_from_sales': total_qty,
                'was_present': True,
                'hours_expected': Decimal('8'),
            },
        )
        if not created:
            DailyReport.objects.filter(pk=report.pk).update(
                units_from_sales=F('units_from_sales') + total_qty
            )


def _sales_redirect(request):
    q = request.POST.get('preserve_query', '').strip()
    if q:
        return redirect(f'/sales/?{q}')
    return redirect('sales_dashboard')


def _money(value):
    if value is None:
        return _decimal_zero()
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        return _decimal_zero()


def _record_accounting_income(sale, payment_method, total_price):
    register_name = REGISTER_BY_PAYMENT.get(payment_method, "Main Cash")
    register, _ = CashRegister.objects.select_for_update().get_or_create(
        name=register_name, defaults={"balance": _decimal_zero()}
    )
    register.balance = _money(register.balance) + total_price
    register.save(update_fields=["balance"])
    AccountingTransaction.objects.create(
        cash_register=register,
        amount=total_price,
        transaction_type="income",
        description=f"Sotuv #{sale.id} ({sale.get_payment_method_display()})",
    )


def _date_range(request):
    period = request.GET.get('period', 'today')
    today = timezone.localdate()
    if period == 'week':
        return today - timedelta(days=6), today
    if period == 'custom':
        try:
            df = date.fromisoformat(request.GET.get('date_from', ''))
            dt = date.fromisoformat(request.GET.get('date_to', ''))
            return df, dt
        except ValueError:
            pass
    return today, today


@login_required
def sales_dashboard(request):
    if not _can_access(request.user, 'seller'):
        return redirect('dashboard')

    date_from, date_to = _date_range(request)

    if request.method == 'POST':
        try:
            if 'save_shift_allocation' in request.POST:
                _save_shift_allocations(request)
            elif 'close_shift' in request.POST:
                _close_shift_session(request)
            elif 'multi_sale' in request.POST:
                _handle_multi_sale(request)
            elif 'quick_sale' in request.POST:
                _handle_quick_sale(request)
            elif 'add_return' in request.POST:
                _handle_return(request)
        except FinishedGoodsInventory.DoesNotExist:
            messages.error(request, "Bu mahsulot uchun tayyor qoldiq topilmadi.")
        except Product.DoesNotExist:
            messages.error(request, "Mahsulot topilmadi.")
        except ValueError as error:
            messages.error(request, str(error))
        return _sales_redirect(request)

    if request.GET.get('export') == 'excel':
        return _export_sales_excel(date_from, date_to)

    return render(request, 'sales.html', _sales_context(date_from, date_to, request))


def _save_shift_allocations(request):
    d = timezone.localdate()
    raw_date = request.POST.get('alloc_date', '').strip()
    if raw_date:
        d = date.fromisoformat(raw_date)
    shift_name = request.POST.get('shift_name', '').strip()
    if not shift_name:
        raise ValueError("Smena tanlang.")
    if ShiftClosure.objects.filter(date=d, shift_name=shift_name).exists():
        raise ValueError("Smena yopilgan — ajratishlarni o'zgartirolmaysiz.")
    product_ids = request.POST.getlist('alloc_product_id')
    quantities = request.POST.getlist('alloc_qty')
    with transaction.atomic():
        ShiftDailyAllocation.objects.filter(date=d, shift_name=shift_name).delete()
        created = 0
        for pid, qty_raw in zip(product_ids, quantities):
            if not pid or not qty_raw:
                continue
            qty = int(qty_raw)
            if qty <= 0:
                continue
            ShiftDailyAllocation.objects.create(
                date=d, shift_name=shift_name, product_id=int(pid), allocated_qty=qty
            )
            created += 1
    if created == 0:
        raise ValueError("Kamida bitta mahsulot uchun ajratilgan miqdor kiriting.")
    messages.success(request, "Smena ajratishlari saqlandi.")


def _close_shift_session(request):
    d = timezone.localdate()
    raw_date = request.POST.get('close_shift_date', '').strip()
    if raw_date:
        d = date.fromisoformat(raw_date)
    shift_name = request.POST.get('shift_name', '').strip()
    if not shift_name:
        raise ValueError("Smena tanlang.")
    notes = request.POST.get('close_notes', '').strip()
    ShiftClosure.objects.update_or_create(
        date=d,
        shift_name=shift_name,
        defaults={'closed_by': request.user, 'notes': notes},
    )
    messages.success(request, f"'{shift_name}' smenasi yopildi.")


def _get_seller_for_shift(request, shift_name):
    seller_user = request.user
    if shift_name:
        shift_emp = Employee.objects.filter(shift__name=shift_name, user_account__isnull=False, status='active').first()
        if shift_emp and shift_emp.user_account:
            seller_user = shift_emp.user_account
    return seller_user


def _handle_multi_sale(request):
    """Handle a single sale with multiple product lines."""
    payment_method = request.POST.get('payment_method', 'cash')
    shift_name = request.POST.get('shift_name', '').strip()
    product_ids = request.POST.getlist('product_id')
    quantities = request.POST.getlist('quantity')

    if not product_ids:
        raise ValueError("Kamida bitta mahsulot tanlang.")

    valid_payments = {c[0] for c in Sale.PAYMENT_CHOICES}
    if payment_method not in valid_payments:
        raise ValueError("To'lov usuli noto'g'ri.")

    with transaction.atomic():
        total = _decimal_zero()
        items_data = []
        for pid, qty_raw in zip(product_ids, quantities):
            if not pid or not qty_raw:
                continue
            qty = int(qty_raw)
            if qty <= 0:
                continue
            inv = FinishedGoodsInventory.objects.select_for_update().select_related('product').get(product_id=pid)
            if inv.stock < qty:
                raise ValueError(f"{inv.product.name}: yetarli emas (bor: {inv.stock}).")
            price = _money(inv.product.price)
            total += price * qty
            items_data.append((inv, qty, price))

        if not items_data:
            raise ValueError("Hech qanday mahsulot tanlanmadi.")

        sale = Sale.objects.create(
            total_amount=total,
            payment_method=payment_method,
            seller=_get_seller_for_shift(request, shift_name),
            shift_name=shift_name,
        )
        for inv, qty, price in items_data:
            SaleItem.objects.create(sale=sale, product=inv.product, quantity=qty, price_at_sale=price)
            inv.stock -= qty
            inv.save(update_fields=['stock'])

        _record_accounting_income(sale, payment_method, total)

    messages.success(request, f"Sotuv #{sale.id} yakunlandi: {total:,.0f} UZS.")
    emp_id = request.POST.get('seller_employee_id')
    _credit_piecework_for_sale(sale, employee_id=emp_id)


def _handle_quick_sale(request):
    product_id = request.POST.get("product")
    qty = int(request.POST.get("quantity", 0))
    payment_method = request.POST.get("payment_method", "cash")
    shift_name = request.POST.get("shift_name", "").strip()

    if qty <= 0:
        raise ValueError("Miqdor 0 dan katta bo'lishi kerak.")

    with transaction.atomic():
        inv = FinishedGoodsInventory.objects.select_for_update().select_related('product').get(product_id=product_id)
        if inv.stock < qty:
            raise ValueError(f"{inv.product.name}: yetarli emas.")
        price = _money(inv.product.price)
        total = price * qty
        sale = Sale.objects.create(
            total_amount=total, payment_method=payment_method,
            seller=_get_seller_for_shift(request, shift_name), shift_name=shift_name,
        )
        SaleItem.objects.create(sale=sale, product=inv.product, quantity=qty, price_at_sale=price)
        inv.stock -= qty
        inv.save(update_fields=['stock'])
        _record_accounting_income(sale, payment_method, total)

    messages.success(request, f"{qty} x {inv.product.name} = {total:,.0f} UZS.")
    emp_id = request.POST.get('seller_employee_id')
    _credit_piecework_for_sale(sale, employee_id=emp_id)


def _handle_return(request):
    product_id = request.POST.get("product")
    qty = int(request.POST.get("quantity", 0))
    reason = request.POST.get("reason")
    if qty <= 0:
        raise ValueError("Miqdor 0 dan katta bo'lishi kerak.")
    product = Product.objects.get(id=product_id)
    ReturnLog.objects.create(product=product, quantity=qty, reason=reason)
    messages.warning(request, f"Qaytarish: {qty} x {product.name}.")


def _sales_context(date_from, date_to, request=None):
    today = timezone.localdate()
    first_of_month = today.replace(day=1)

    categories = ProductCategory.objects.prefetch_related('products').all()
    products = list(Product.objects.select_related('category').order_by('name'))
    finished_goods_qs = FinishedGoodsInventory.objects.select_related('product__category').order_by('product__name')
    finished_goods_all = list(finished_goods_qs)

    inv_map = {fg.product_id: fg for fg in finished_goods_all}
    product_options = [
        {
            'product': p,
            'stock': inv_map[p.id].stock if p.id in inv_map else 0,
            'has_inventory': p.id in inv_map,
        }
        for p in products
    ]

    sale_cart_products = [
        {
            'id': p.id,
            'name': p.name,
            'price': float(p.price),
            'category': p.category.name if p.category else 'Turkumsiz',
            'stock': inv_map[p.id].stock if p.id in inv_map else 0,
            'disabled': p.id not in inv_map or inv_map[p.id].stock <= 0,
            'image_url': p.image.url if p.image else '',
        }
        for p in products
    ]

    today_sales = Sale.objects.filter(date__date=today)
    today_revenue = today_sales.aggregate(t=Sum('total_amount'))['t'] or _decimal_zero()
    today_count = today_sales.count()

    month_sales = Sale.objects.filter(date__date__gte=first_of_month)
    month_revenue = month_sales.aggregate(t=Sum('total_amount'))['t'] or _decimal_zero()
    month_sales_count = month_sales.count()

    _line_val = ExpressionWrapper(F('quantity') * F('price_at_sale'), output_field=DecimalField(max_digits=14, decimal_places=2))
    top_today = list(
        SaleItem.objects.filter(sale__date__date=today)
        .annotate(lv=_line_val)
        .values('product__name')
        .annotate(quantity=Sum('quantity'), revenue=Sum('lv'))
        .order_by('-revenue')[:5]
    )
    top_products_all = list(
        SaleItem.objects.filter(sale__date__date__gte=first_of_month)
        .values('product__name')
        .annotate(total_qty=Sum('quantity'))
        .order_by('-total_qty')[:5]
    )

    recent_sales = (
        Sale.objects
        .filter(date__date__gte=date_from, date__date__lte=date_to)
        .prefetch_related('items__product__category')
        .select_related('seller')
        .order_by('-date')[:50]
    )

    last_7 = []
    for offset in range(6, -1, -1):
        d = today - timedelta(days=offset)
        rev = Sale.objects.filter(date__date=d).aggregate(t=Sum('total_amount'))['t'] or _decimal_zero()
        last_7.append({'date': d.strftime('%d.%m'), 'revenue': float(rev)})

    total_inv_value = sum(_money(fg.product.price) * fg.stock for fg in finished_goods_all)
    low_stock = [fg for fg in finished_goods_all if 0 < fg.stock < 10]
    out_of_stock = [fg for fg in finished_goods_all if fg.stock == 0]

    # Returns stats
    from .models import ReturnLog
    today_returns = ReturnLog.objects.filter(date__date=today)
    returns_count = today_returns.count()
    returns_quantity = today_returns.aggregate(t=Sum('quantity'))['t'] or 0
    returns_brak = today_returns.filter(reason='brak').aggregate(t=Sum('quantity'))['t'] or 0
    returns_unsold = today_returns.filter(reason='unsold').aggregate(t=Sum('quantity'))['t'] or 0

    # Payment breakdown today
    payment_breakdown = list(
        today_sales.values('payment_method')
        .annotate(count=Count('id'), total=Sum('total_amount'))
        .order_by('payment_method')
    )

    all_sales_count = Sale.objects.count()
    avg_sale_value = Sale.objects.aggregate(avg=Sum('total_amount'))['avg'] or _decimal_zero()
    if all_sales_count:
        avg_sale_value = avg_sale_value / all_sales_count

    today_items_sold = SaleItem.objects.filter(sale__date__date=today).aggregate(t=Sum('quantity'))['t'] or 0
    month_items = SaleItem.objects.filter(sale__date__date__gte=first_of_month).aggregate(t=Sum('quantity'))['t'] or 0

    from hr.models import Employee, Shift
    shifts = list(Shift.objects.all().order_by('name'))

    focus_shift = ''
    emp_shift_name = ''
    if request and request.user.is_authenticated:
        try:
            emp = Employee.objects.select_related('shift').get(user_account=request.user)
            if emp.shift:
                emp_shift_name = emp.shift.name
        except Employee.DoesNotExist:
            pass

    alloc_focus_day = today
    preserve_query = ''
    if request:
        preserve_query = request.GET.urlencode()
        focus_shift = request.GET.get('shift', '').strip()
        raw_alloc_date = request.GET.get('alloc_date', '').strip()
        if raw_alloc_date:
            try:
                alloc_focus_day = date.fromisoformat(raw_alloc_date)
            except ValueError:
                pass
    if not focus_shift:
        if emp_shift_name:
            focus_shift = emp_shift_name
        elif shifts:
            focus_shift = shifts[0].name

    shift_remainder_rows = []
    shift_closed_focus = False
    if focus_shift:
        shift_closed_focus = ShiftClosure.objects.filter(
            date=alloc_focus_day, shift_name=focus_shift
        ).exists()
        for alloc in ShiftDailyAllocation.objects.filter(
            date=alloc_focus_day, shift_name=focus_shift
        ).select_related('product'):
            sold = _sold_qty_shift_product(alloc_focus_day, focus_shift, alloc.product_id)
            shift_remainder_rows.append({
                'product_name': alloc.product.name,
                'allocated': alloc.allocated_qty,
                'sold': sold,
                'remainder': max(alloc.allocated_qty - sold, 0),
            })

    shift_sales_breakdown = list(
        Sale.objects.filter(date__date__gte=date_from, date__date__lte=date_to)
        .exclude(shift_name='')
        .values('shift_name')
        .annotate(revenue=Sum('total_amount'), checks=Count('id'))
        .order_by('shift_name')
    )

    period_revenue = (
        Sale.objects.filter(date__date__gte=date_from, date__date__lte=date_to).aggregate(
            t=Sum('total_amount')
        )['t']
        or _decimal_zero()
    )

    return {
        'products': products,
        'product_options': product_options,
        'categories': categories,
        'employees': Employee.objects.filter(status='active', is_piecework=True).select_related('shift').order_by('name'),
        'finished_goods': [
            {
                'product': fg.product,
                'stock': fg.stock,
                'item_value': _money(fg.product.price) * fg.stock,
                'fg_id': fg.id,
                'produced_at': fg.produced_at,
            }
            for fg in finished_goods_all
        ],
        'finished_goods_raw': finished_goods_qs,
        'recent_sales': recent_sales,
        'today_sales_count': today_count,
        'today_revenue': today_revenue,
        'month_revenue': month_revenue,
        'month_sales_count': month_sales_count,
        'month_items': month_items,
        'today_items_sold': today_items_sold,
        'top_products_today': top_today,
        'top_products_all': top_products_all,
        'last_7_days': last_7,
        'low_stock_items': low_stock,
        'out_of_stock_items': out_of_stock,
        'total_inventory_value': total_inv_value,
        'returns_count': returns_count,
        'returns_quantity': returns_quantity,
        'returns_brak': returns_brak,
        'returns_unsold': returns_unsold,
        'payment_breakdown': payment_breakdown,
        'all_sales_count': all_sales_count,
        'avg_sale_value': avg_sale_value,
        'shifts': shifts,
        'date_from': date_from,
        'date_to': date_to,
        'period': request.GET.get('period', 'today') if request else 'today',
        'focus_shift': focus_shift,
        'alloc_focus_day': alloc_focus_day,
        'shift_remainder_rows': shift_remainder_rows,
        'shift_closed_focus': shift_closed_focus,
        'shift_sales_breakdown': shift_sales_breakdown,
        'period_revenue': period_revenue,
        'preserve_query': preserve_query,
        'sale_cart_products': sale_cart_products,
    }


def _export_sales_excel(date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sotuvlar"
    fill = PatternFill("solid", fgColor="D4A373")
    bold = Font(bold=True)
    headers = ['Sana', 'Chek #', 'Mahsulot', 'Kategoriya', 'Miqdor', 'Narx', 'Jami', 'To\'lov', 'Smena', 'Sotuvchi']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    sales = Sale.objects.filter(
        date__date__gte=date_from, date__date__lte=date_to
    ).prefetch_related('items__product__category').select_related('seller').order_by('-date')

    row_idx = 2
    for sale in sales:
        for item in sale.items.all():
            ws.cell(row=row_idx, column=1, value=sale.date.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row_idx, column=2, value=sale.id)
            ws.cell(row=row_idx, column=3, value=item.product.name)
            ws.cell(row=row_idx, column=4, value=item.product.category.name if item.product.category else '')
            ws.cell(row=row_idx, column=5, value=item.quantity)
            ws.cell(row=row_idx, column=6, value=float(item.price_at_sale))
            ws.cell(row=row_idx, column=7, value=float(item.quantity * item.price_at_sale))
            ws.cell(row=row_idx, column=8, value=sale.get_payment_method_display())
            ws.cell(row=row_idx, column=9, value=sale.shift_name or '')
            ws.cell(row=row_idx, column=10, value=sale.seller.username if sale.seller else '')
            row_idx += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="sales_{date_from}_{date_to}.xlsx"'
    return resp
