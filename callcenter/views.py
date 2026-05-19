import json
import hashlib
import hmac
import io
from datetime import timedelta

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator

from .models import (
    MetaLead, LeadStatus, LeadStatusHistory, LeadComment,
    Callback, Notification, ApiLog, ReportLog, SystemSettings
)
from .services import process_meta_lead, create_test_lead, assign_operator_round_robin


# ── Role helpers ──────────────────────────────────────────────────────

def _is_cc_staff(user):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=['callcenter_operator', 'callcenter_supervisor', 'callcenter_admin']).exists()


def _is_cc_admin(user):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=['callcenter_admin', 'callcenter_supervisor']).exists()


def cc_required(view):
    return user_passes_test(_is_cc_staff, login_url='login')(login_required(view))


def cc_admin_required(view):
    return user_passes_test(_is_cc_admin, login_url='login')(login_required(view))


def _ensure_cc_groups():
    for name in ('callcenter_operator', 'callcenter_supervisor', 'callcenter_admin'):
        Group.objects.get_or_create(name=name)


# ── Dashboard ─────────────────────────────────────────────────────────

@login_required
@cc_required
def dashboard(request):
    _ensure_cc_groups()
    today = timezone.now().date()

    qs = MetaLead.objects.all()
    if not _is_cc_admin(request.user):
        qs = qs.filter(assigned_operator=request.user)

    today_leads = qs.filter(crm_received_time__date=today).count()
    today_called = qs.filter(
        crm_received_time__date=today,
        current_status__code='called'
    ).count()
    today_not_called = qs.filter(
        crm_received_time__date=today,
        current_status__code='not_called'
    ).count()
    today_received = qs.filter(
        crm_received_time__date=today,
        current_status__code='received'
    ).count()
    pending_callback = qs.filter(current_status__code='callback', is_closed=False).count()
    total = qs.filter(crm_received_time__date=today).count() or 1
    conversion_rate = round((today_received / total) * 100, 1)

    # Chart: last 7 days lead flow
    chart_labels = []
    chart_data = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        chart_labels.append(d.strftime('%d %b'))
        chart_data.append(qs.filter(crm_received_time__date=d).count())

    # Status distribution
    status_dist = list(
        qs.values('current_status__name', 'current_status__color')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )

    # Operator ranking (admin only)
    operator_ranking = []
    if _is_cc_admin(request.user):
        operator_ranking = list(
            MetaLead.objects.filter(crm_received_time__date=today)
            .values('assigned_operator__username')
            .annotate(cnt=Count('id'))
            .order_by('-cnt')[:10]
        )

    new_leads = qs.filter(current_status__code='new').order_by('-crm_received_time')[:10]
    unread_notifications = request.user.cc_notifications.filter(is_read=False).count()

    context = {
        'today_leads': today_leads,
        'today_called': today_called,
        'today_not_called': today_not_called,
        'today_received': today_received,
        'pending_callback': pending_callback,
        'conversion_rate': conversion_rate,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'status_dist': status_dist,
        'operator_ranking': operator_ranking,
        'new_leads': new_leads,
        'unread_notifications': unread_notifications,
        'page_title': 'Call Center Dashboard',
        'is_cc_admin': _is_cc_admin(request.user),
    }
    return render(request, 'callcenter/dashboard.html', context)


# ── Leads ─────────────────────────────────────────────────────────────

@login_required
@cc_required
def lead_list(request):
    qs = MetaLead.objects.select_related('current_status', 'assigned_operator').all()

    if not _is_cc_admin(request.user):
        qs = qs.filter(assigned_operator=request.user)

    search = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    operator_filter = request.GET.get('operator', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    region_filter = request.GET.get('region', '')
    campaign_filter = request.GET.get('campaign', '')

    if search:
        qs = qs.filter(
            Q(full_name__icontains=search) |
            Q(phone_number__icontains=search) |
            Q(phone_number2__icontains=search)
        )
    if status_filter:
        qs = qs.filter(current_status__code=status_filter)
    if operator_filter and _is_cc_admin(request.user):
        qs = qs.filter(assigned_operator_id=operator_filter)
    if date_from:
        qs = qs.filter(crm_received_time__date__gte=date_from)
    if date_to:
        qs = qs.filter(crm_received_time__date__lte=date_to)
    if region_filter:
        qs = qs.filter(region__icontains=region_filter)
    if campaign_filter:
        qs = qs.filter(campaign_name__icontains=campaign_filter)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    statuses = LeadStatus.objects.all()
    operators = get_operators()

    context = {
        'page_obj': page_obj,
        'statuses': statuses,
        'operators': operators,
        'search': search,
        'status_filter': status_filter,
        'operator_filter': operator_filter,
        'date_from': date_from,
        'date_to': date_to,
        'region_filter': region_filter,
        'campaign_filter': campaign_filter,
        'page_title': 'Leadlar',
        'is_cc_admin': _is_cc_admin(request.user),
    }
    return render(request, 'callcenter/lead_list.html', context)


@login_required
@cc_required
def lead_detail(request, pk):
    lead = get_object_or_404(
        MetaLead.objects.select_related('current_status', 'assigned_operator'),
        pk=pk
    )
    if not _is_cc_admin(request.user) and lead.assigned_operator != request.user:
        messages.error(request, 'Ruxsat yo\'q')
        return redirect('cc_lead_list')

    history = lead.history.select_related('old_status', 'new_status', 'operator').all()
    comments = lead.comments.select_related('operator').all()
    callbacks = lead.callbacks.all()
    statuses = LeadStatus.objects.all()
    operators = get_operators()

    context = {
        'lead': lead,
        'history': history,
        'comments': comments,
        'callbacks': callbacks,
        'statuses': statuses,
        'operators': operators,
        'page_title': f'Lead: {lead.full_name}',
        'is_cc_admin': _is_cc_admin(request.user),
    }
    return render(request, 'callcenter/lead_detail.html', context)


@login_required
@cc_required
@require_POST
def lead_update_status(request, pk):
    lead = get_object_or_404(MetaLead, pk=pk)
    if not _is_cc_admin(request.user) and lead.assigned_operator != request.user:
        return JsonResponse({'error': 'Ruxsat yo\'q'}, status=403)

    status_code = request.POST.get('status_code')
    comment = request.POST.get('comment', '').strip()
    callback_date = request.POST.get('callback_date', '')

    new_status = get_object_or_404(LeadStatus, code=status_code)

    # Business rules
    if status_code == 'called' and not comment:
        return JsonResponse({'error': 'Qo\'ngiroq natijasi uchun izoh majburiy'}, status=400)
    if status_code == 'callback' and not callback_date:
        return JsonResponse({'error': 'Callback sanasi majburiy'}, status=400)

    old_status = lead.current_status
    lead.current_status = new_status
    if callback_date:
        lead.callback_date = callback_date
    if status_code in ('received', 'cancelled', 'wrong_number'):
        lead.is_closed = True
    lead.save()

    LeadStatusHistory.objects.create(
        lead=lead,
        old_status=old_status,
        new_status=new_status,
        operator=request.user,
        comment=comment,
    )

    if status_code == 'callback' and callback_date:
        Callback.objects.create(
            lead=lead,
            callback_datetime=callback_date,
            callback_note=comment,
            created_by=request.user,
        )

    if status_code == 'not_called':
        Notification.objects.create(
            user=request.user,
            lead=lead,
            notification_text=f"Ertaga eslatma: {lead.full_name} ga qo'ngiroq qiling",
            notification_type='reminder'
        )

    return JsonResponse({'status': 'ok', 'new_status': new_status.name, 'color': new_status.color})


@login_required
@cc_required
@require_POST
def lead_add_comment(request, pk):
    lead = get_object_or_404(MetaLead, pk=pk)
    text = request.POST.get('comment_text', '').strip()
    if not text:
        return JsonResponse({'error': 'Izoh bo\'sh bo\'lmasin'}, status=400)

    comment = LeadComment.objects.create(
        lead=lead,
        operator=request.user,
        comment_text=text,
    )
    return JsonResponse({
        'status': 'ok',
        'comment': text,
        'operator': request.user.get_full_name() or request.user.username,
        'created_at': comment.created_at.strftime('%d.%m.%Y %H:%M'),
    })


@login_required
@cc_required
@require_POST
def lead_assign(request, pk):
    if not _is_cc_admin(request.user):
        return JsonResponse({'error': 'Ruxsat yo\'q'}, status=403)
    lead = get_object_or_404(MetaLead, pk=pk)
    operator_id = request.POST.get('operator_id')
    operator = get_object_or_404(User, pk=operator_id)
    lead.assigned_operator = operator
    lead.save(update_fields=['assigned_operator'])
    return JsonResponse({'status': 'ok', 'operator': operator.username})


# ── Notifications ─────────────────────────────────────────────────────

@login_required
@cc_required
def notifications(request):
    notifs = request.user.cc_notifications.select_related('lead').all()
    request.user.cc_notifications.filter(is_read=False).update(is_read=True)

    paginator = Paginator(notifs, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj': page_obj,
        'page_title': 'Bildirishnomalar',
    }
    return render(request, 'callcenter/notifications.html', context)


@login_required
@cc_required
@require_GET
def notifications_count(request):
    count = request.user.cc_notifications.filter(is_read=False).count()
    return JsonResponse({'count': count})


# ── Reports ───────────────────────────────────────────────────────────

@login_required
@cc_required
def reports(request):
    today = timezone.now().date()
    date_from = request.GET.get('date_from', (today - timedelta(days=30)).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())
    operator_filter = request.GET.get('operator', '')
    region_filter = request.GET.get('region', '')
    campaign_filter = request.GET.get('campaign', '')

    qs = MetaLead.objects.all()
    if not _is_cc_admin(request.user):
        qs = qs.filter(assigned_operator=request.user)
    if date_from:
        qs = qs.filter(crm_received_time__date__gte=date_from)
    if date_to:
        qs = qs.filter(crm_received_time__date__lte=date_to)
    if operator_filter and _is_cc_admin(request.user):
        qs = qs.filter(assigned_operator_id=operator_filter)
    if region_filter:
        qs = qs.filter(region__icontains=region_filter)
    if campaign_filter:
        qs = qs.filter(campaign_name__icontains=campaign_filter)

    total = qs.count()
    received = qs.filter(current_status__code='received').count()
    conversion = round((received / total * 100), 1) if total else 0

    status_dist = list(
        qs.values('current_status__name', 'current_status__color')
        .annotate(cnt=Count('id')).order_by('-cnt')
    )

    operator_stats = list(
        qs.values('assigned_operator__username')
        .annotate(
            total=Count('id'),
            received=Count('id', filter=Q(current_status__code='received')),
            called=Count('id', filter=Q(current_status__code='called')),
        ).order_by('-total')
    )

    # Daily chart
    from django.db.models.functions import TruncDate
    daily = list(
        qs.annotate(day=TruncDate('crm_received_time'))
        .values('day').annotate(cnt=Count('id')).order_by('day')
    )
    daily_labels = json.dumps([str(d['day']) for d in daily])
    daily_data = json.dumps([d['cnt'] for d in daily])

    ReportLog.objects.create(report_type='leads_report', generated_by=request.user)

    context = {
        'total': total,
        'received': received,
        'conversion': conversion,
        'status_dist': status_dist,
        'operator_stats': operator_stats,
        'daily_labels': daily_labels,
        'daily_data': daily_data,
        'date_from': date_from,
        'date_to': date_to,
        'operator_filter': operator_filter,
        'region_filter': region_filter,
        'campaign_filter': campaign_filter,
        'operators': get_operators(),
        'page_title': 'Hisobotlar',
        'is_cc_admin': _is_cc_admin(request.user),
    }
    return render(request, 'callcenter/reports.html', context)


@login_required
@cc_required
def export_leads_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    qs = MetaLead.objects.select_related('current_status', 'assigned_operator').all()
    if not _is_cc_admin(request.user):
        qs = qs.filter(assigned_operator=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads'

    headers = ['#', 'Ism', 'Telefon', 'Telefon 2', 'Hudud', 'Mahsulot', 'Kampaniya',
               'Status', 'Operator', 'Callback', 'Yopilgan', 'Sana']
    header_fill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, lead in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=lead.full_name)
        ws.cell(row=row_idx, column=3, value=lead.phone_number)
        ws.cell(row=row_idx, column=4, value=lead.phone_number2)
        ws.cell(row=row_idx, column=5, value=lead.region)
        ws.cell(row=row_idx, column=6, value=lead.product_interest)
        ws.cell(row=row_idx, column=7, value=lead.campaign_name)
        ws.cell(row=row_idx, column=8, value=lead.current_status.name if lead.current_status else '')
        ws.cell(row=row_idx, column=9, value=lead.assigned_operator.username if lead.assigned_operator else '')
        ws.cell(row=row_idx, column=10, value=str(lead.callback_date) if lead.callback_date else '')
        ws.cell(row=row_idx, column=11, value='Ha' if lead.is_closed else 'Yo\'q')
        ws.cell(row=row_idx, column=12, value=lead.crm_received_time.strftime('%d.%m.%Y %H:%M'))

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ReportLog.objects.create(report_type='excel_export', generated_by=request.user)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="leads_export.xlsx"'
    return response


# ── User Management ───────────────────────────────────────────────────

@login_required
@cc_admin_required
def user_management(request):
    _ensure_cc_groups()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            full_name = request.POST.get('full_name', '').strip()
            phone = request.POST.get('phone', '').strip()
            role = request.POST.get('role', 'callcenter_operator')

            if User.objects.filter(username=username).exists():
                messages.error(request, 'Bu username band')
            elif username and password:
                u = User.objects.create_user(username=username, password=password)
                names = full_name.split(' ', 1)
                u.first_name = names[0]
                u.last_name = names[1] if len(names) > 1 else ''
                u.save()
                group, _ = Group.objects.get_or_create(name=role)
                u.groups.add(group)
                messages.success(request, f'Foydalanuvchi yaratildi: {username}')
            else:
                messages.error(request, 'Username va parol majburiy')

        elif action == 'toggle_active':
            uid = request.POST.get('user_id')
            u = get_object_or_404(User, pk=uid)
            if u != request.user:
                u.is_active = not u.is_active
                u.save()
                messages.success(request, f'{u.username} holati o\'zgartirildi')

        elif action == 'delete':
            uid = request.POST.get('user_id')
            u = get_object_or_404(User, pk=uid)
            if u != request.user:
                u.delete()
                messages.success(request, 'Foydalanuvchi o\'chirildi')

        return redirect('cc_users')

    users = User.objects.filter(
        groups__name__in=['callcenter_operator', 'callcenter_supervisor', 'callcenter_admin']
    ).distinct().prefetch_related('groups')

    context = {
        'users': users,
        'page_title': 'Foydalanuvchilar',
    }
    return render(request, 'callcenter/user_management.html', context)


# ── Settings ──────────────────────────────────────────────────────────

@login_required
@cc_admin_required
def system_settings(request):
    cfg = SystemSettings.get()
    if request.method == 'POST':
        cfg.assign_mode = request.POST.get('assign_mode', 'round_robin')
        cfg.meta_verify_token = request.POST.get('meta_verify_token', '').strip()
        cfg.meta_access_token = request.POST.get('meta_access_token', '').strip()
        cfg.meta_app_secret = request.POST.get('meta_app_secret', '').strip()
        cfg.meta_page_id = request.POST.get('meta_page_id', '').strip()
        cfg.save()
        messages.success(request, 'Sozlamalar saqlandi')
        return redirect('cc_settings')

    context = {
        'cfg': cfg,
        'page_title': 'Tizim sozlamalari',
    }
    return render(request, 'callcenter/settings.html', context)


# ── Meta Webhook ──────────────────────────────────────────────────────

@csrf_exempt
def meta_webhook(request):
    cfg = SystemSettings.get()

    if request.method == 'GET':
        mode = request.GET.get('hub.mode')
        token = request.GET.get('hub.verify_token')
        challenge = request.GET.get('hub.challenge')
        verify_token = cfg.meta_verify_token or getattr(settings, 'META_VERIFY_TOKEN', '')
        if mode == 'subscribe' and token == verify_token:
            return HttpResponse(challenge, content_type='text/plain')
        return HttpResponse('Forbidden', status=403)

    if request.method == 'POST':
        try:
            body = request.body
            data = json.loads(body)

            ApiLog.objects.create(
                api_name='meta_webhook_incoming',
                request_payload=data,
                status_code=200,
            )

            for entry in data.get('entry', []):
                for change in entry.get('changes', []):
                    if change.get('field') == 'leadgen':
                        value = change.get('value', {})
                        lead_id = value.get('leadgen_id')
                        page_id = value.get('page_id')
                        form_id = value.get('form_id')
                        if lead_id:
                            process_meta_lead(lead_id, page_id=page_id, form_id=form_id, raw_payload=value)

        except Exception as e:
            ApiLog.objects.create(
                api_name='meta_webhook_error',
                request_payload={'error': str(e)},
                status_code=500,
            )
        return HttpResponse('EVENT_RECEIVED', content_type='text/plain')

    return HttpResponse('Method not allowed', status=405)


# ── Test Lead (no real Meta API needed) ───────────────────────────────

@login_required
@cc_admin_required
@require_POST
def create_test_lead_view(request):
    lead = create_test_lead(
        full_name=request.POST.get('full_name', 'Test Foydalanuvchi'),
        phone=request.POST.get('phone', '+998901234567'),
        region=request.POST.get('region', 'Toshkent'),
        product_interest=request.POST.get('product_interest', 'Test mahsulot'),
        campaign=request.POST.get('campaign', 'Test Campaign'),
    )
    messages.success(request, f'Test lead yaratildi: {lead.full_name}')
    return redirect('cc_lead_list')


# ── API: fetch lead manually ──────────────────────────────────────────

@login_required
@cc_admin_required
@require_POST
def api_fetch_lead(request):
    lead_id = request.POST.get('lead_id', '').strip()
    if not lead_id:
        return JsonResponse({'error': 'lead_id majburiy'}, status=400)
    lead, status = process_meta_lead(lead_id)
    if status == 'created':
        return JsonResponse({'status': 'ok', 'lead_id': lead.pk, 'name': lead.full_name})
    return JsonResponse({'status': status})


# ── Helpers ───────────────────────────────────────────────────────────

def get_operators():
    return User.objects.filter(
        is_active=True,
        groups__name__in=['callcenter_operator', 'callcenter_supervisor', 'callcenter_admin']
    ).distinct()
