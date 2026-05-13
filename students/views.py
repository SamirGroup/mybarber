from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg, F
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings
import json
import io
import os
from datetime import date, timedelta

from .models import (
    Classroom, Student, Parent, DocumentType, StudentDocument,
    Subject, Schedule, Quarter, DailyGrade, QuarterGrade,
    Attendance, Homework, Contract, Payment, SmsNotificationConfig, SmsLog,
    ChatGroup, ChatMessage, HomeworkSubmission, StudentTransfer, StudentBalance,
    OnlinePayment, LessonPeriod
)
from enrollment.models import Lead, Grade, AcademicYear
from django.contrib.auth.models import User, Group


# ── Role helpers ──────────────────────────────────────────────────────
def _is_students_staff(user):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=['students_agent', 'students_manager', 'enrollment_agent', 'enrollment_manager']).exists()


def students_required(view):
    return user_passes_test(_is_students_staff, login_url='login')(login_required(view))


def _ensure_students_groups():
    for name in ('students_agent', 'students_manager'):
        Group.objects.get_or_create(name=name)


# ── Dashboard ─────────────────────────────────────────────────────────
@login_required
@students_required
def dashboard(request):
    _ensure_students_groups()

    # Stats
    students_total = Student.objects.count()
    students_active = Student.objects.filter(is_active=True).count()
    students_in_classroom = Student.objects.exclude(classroom__isnull=True).count()

    classrooms_total = Classroom.objects.count()
    subjects_total = Subject.objects.count()
    contracts_total = Contract.objects.filter(is_active=True).count()

    # Debtors
    today = timezone.now().date()
    debtors = []
    for contract in Contract.objects.filter(is_active=True):
        paid = contract.payments.filter(payment_date__month=today.month, payment_date__year=today.year).aggregate(s=Sum('amount'))['s'] or 0
        if paid < contract.effective_fee:
            debtors.append({
                'contract': contract,
                'debt': contract.effective_fee - paid,
                'parents': contract.student.parents.all()
            })

    # Recent activities
    recent_students = Student.objects.select_related('classroom').order_by('-created_at')[:10]
    recent_payments = Payment.objects.select_related('student', 'contract').order_by('-payment_date')[:10]
    recent_homework = Homework.objects.select_related('classroom', 'subject').order_by('-due_date')[:5]

    # Upcoming birthdays
    today = timezone.now().date()
    upcoming_birthdays = Student.objects.filter(
        birth_date__month=today.month, birth_date__day=today.day
    )[:5]

    context = {
        'students_total': students_total,
        'students_active': students_active,
        'students_in_classroom': students_in_classroom,
        'classrooms_total': classrooms_total,
        'subjects_total': subjects_total,
        'contracts_total': contracts_total,
        'debtors': debtors,
        'recent_students': recent_students,
        'recent_payments': recent_payments,
        'recent_homework': recent_homework,
        'upcoming_birthdays': upcoming_birthdays,
        'page_title': 'O\'quvchilar bo\'limi',
    }
    return render(request, 'students/dashboard.html', context)


# ── Classrooms ────────────────────────────────────────────────────────
@login_required
@students_required
def classroom_list(request):
    year_id = request.GET.get('year')
    classrooms = Classroom.objects.select_related('grade', 'academic_year', 'homeroom_teacher').all()
    
    if year_id:
        classrooms = classrooms.filter(academic_year_id=year_id)
    
    academic_years = AcademicYear.objects.all()
    
    context = {
        'classrooms': classrooms,
        'academic_years': academic_years,
        'page_title': 'Sinf xonalari',
    }
    return render(request, 'students/classroom_list.html', context)


@login_required
@students_required
def classroom_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        grade_id = request.POST.get('grade')
        academic_year_id = request.POST.get('academic_year')
        homeroom_teacher_id = request.POST.get('homeroom_teacher') or None
        capacity = request.POST.get('capacity', 30)
        
        classroom = Classroom(
            name=name,
            grade_id=grade_id,
            academic_year_id=academic_year_id,
            homeroom_teacher_id=homeroom_teacher_id,
            capacity=capacity
        )
        classroom.save()
        
        messages.success(request, f'Sinf yaratildi: {classroom.name}')
        return redirect('students_classroom_list')
    
    grades = Grade.objects.all()
    academic_years = AcademicYear.objects.all()
    teachers = User.objects.filter(groups__name__in=['students_agent', 'students_manager', 'enrollment_agent', 'enrollment_manager']).distinct()
    
    context = {
        'grades': grades,
        'academic_years': academic_years,
        'teachers': teachers,
        'page_title': 'Yangi sinf yaratish',
    }
    return render(request, 'students/classroom_form.html', context)


@login_required
@students_required
def classroom_detail(request, pk):
    classroom = get_object_or_404(
        Classroom.objects.select_related('grade', 'academic_year', 'homeroom_teacher'),
        pk=pk
    )
    students = classroom.students.select_related('classroom').all()
    schedules = classroom.schedules.select_related('subject', 'teacher').all()
    homeworks = classroom.homeworks.select_related('subject', 'teacher').all()
    
    context = {
        'classroom': classroom,
        'students': students,
        'schedules': schedules,
        'homeworks': homeworks,
        'page_title': f'Sinf: {classroom.name}',
    }
    return render(request, 'students/classroom_detail.html', context)


# ── Students ──────────────────────────────────────────────────────────
@login_required
@students_required
def student_list(request):
    classroom_filter = request.GET.get('classroom')
    search = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', 'last_name')
    
    students = Student.objects.select_related('classroom').all()
    
    if classroom_filter:
        students = students.filter(classroom_id=classroom_filter)
    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(passport_or_id__icontains=search) |
            Q(erp_number__icontains=search)
        )
    
    ALLOWED_SORT = {'last_name', 'first_name', 'classroom', '-created_at'}
    sort_by = sort_by if sort_by in ALLOWED_SORT else 'last_name'
    students = students.order_by(sort_by)
    
    paginator = Paginator(students, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    classrooms = Classroom.objects.all()
    
    context = {
        'page_obj': page_obj,
        'classrooms': classrooms,
        'classroom_filter': classroom_filter,
        'search': search,
        'sort_by': sort_by,
        'page_title': 'O\'quvchilar ro\'yxati',
    }
    return render(request, 'students/student_list.html', context)


@login_required
@students_required
def student_create(request):
    if request.method == 'POST':
        student = Student(
            first_name=request.POST.get('first_name', '').strip(),
            last_name=request.POST.get('last_name', '').strip(),
            middle_name=request.POST.get('middle_name', '').strip(),
            gender=request.POST.get('gender', 'M'),
            birth_date=request.POST.get('birth_date') or timezone.now().date(),
            passport_or_id=request.POST.get('passport_or_id', '').strip(),
            erp_number=request.POST.get('erp_number', '').strip(),
            address=request.POST.get('address', '').strip(),
            medical_notes=request.POST.get('medical_notes', '').strip(),
            is_active=request.POST.get('is_active') == 'on',
            classroom_id=request.POST.get('classroom') or None,
        )
        student.save()
        
        # Parents
        parent_ids = request.POST.getlist('parents')
        if parent_ids:
            student.parents.add(*parent_ids)
        
        messages.success(request, f'O\'quvchi yaratildi: {student.full_name}')
        return redirect('students_detail', pk=student.pk)
    
    classrooms = Classroom.objects.all()
    parents = Parent.objects.all()
    
    context = {
        'classrooms': classrooms,
        'parents': parents,
        'page_title': 'Yangi o\'quvchi',
    }
    return render(request, 'students/student_form.html', context)


@login_required
@students_required
def student_detail(request, pk):
    student = get_object_or_404(
        Student.objects.select_related('classroom'),
        pk=pk
    )
    
    # Documents
    documents = student.documents.select_related('doc_type', 'uploaded_by').all()
    doc_types = DocumentType.objects.all()
    
    # Grades
    daily_grades = student.daily_grades.select_related('subject', 'teacher').order_by('-date')[:20]
    quarter_grades = student.quarter_grades.select_related('subject', 'quarter', 'teacher').all()
    
    # Attendance
    attendance = student.attendances.select_related('subject', 'marked_by').order_by('-date')[:30]
    
    # Finance
    contracts = student.contracts.select_related().all()
    payments = student.payments.select_related('contract', 'received_by').order_by('-payment_date')[:20]
    
    # Homework
    homeworks = Homework.objects.filter(classroom=student.classroom).select_related('subject', 'teacher').order_by('-due_date')[:10]
    
    # Chat groups
    chat_groups = student.chat_groups.all()
    
    # Debt calculation
    today = timezone.now().date()
    total_debt = 0
    for contract in contracts.filter(is_active=True):
        paid = contract.payments.filter(
            payment_date__month=today.month,
            payment_date__year=today.year
        ).aggregate(s=Sum('amount'))['s'] or 0
        total_debt += contract.effective_fee - paid
    
    # Quarterly results
    quarters = Quarter.objects.filter(academic_year=student.classroom.academic_year if student.classroom else None)
    quarterly_results = []
    for quarter in quarters:
        grades = student.get_quarter_grades_by_subject(quarter)
        quarterly_results.append({
            'quarter': quarter,
            'grades': grades,
            'average': round(sum(grades.values()) / len(grades), 2) if grades else 0,
        })
    
    # Balance for current month
    balance_info = student.get_balance_for_month(today.year, today.month)
    
    context = {
        'student': student,
        'documents': documents,
        'doc_types': doc_types,
        'daily_grades': daily_grades,
        'quarter_grades': quarter_grades,
        'attendance': attendance,
        'contracts': contracts,
        'payments': payments,
        'homeworks': homeworks,
        'chat_groups': chat_groups,
        'total_debt': total_debt,
        'quarterly_results': quarterly_results,
        'balance_info': balance_info,
        'discount_details': student.get_discount_details(),
        'has_discount': student.has_discount,
        'average_grade': student.get_average_grade(),
        'attendance_rate': student.get_attendance_rate(),
        'page_title': f'O\'quvchi: {student.full_name}',
    }
    return render(request, 'students/student_detail.html', context)


@login_required
@students_required
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        student.first_name = request.POST.get('first_name', '').strip()
        student.last_name = request.POST.get('last_name', '').strip()
        student.middle_name = request.POST.get('middle_name', '').strip()
        student.gender = request.POST.get('gender', 'M')
        student.birth_date = request.POST.get('birth_date') or student.birth_date
        student.passport_or_id = request.POST.get('passport_or_id', '').strip()
        student.erp_number = request.POST.get('erp_number', '').strip()
        student.address = request.POST.get('address', '').strip()
        student.medical_notes = request.POST.get('medical_notes', '').strip()
        student.is_active = request.POST.get('is_active') == 'on'
        student.classroom_id = request.POST.get('classroom') or None
        student.save()
        
        # Update parents
        parent_ids = request.POST.getlist('parents')
        student.parents.set(parent_ids)
        
        messages.success(request, f'O\'quvchi ma\'lumotlari yangilandi: {student.full_name}')
        return redirect('students_detail', pk=student.pk)
    
    classrooms = Classroom.objects.all()
    parents = Parent.objects.all()
    
    context = {
        'student': student,
        'classrooms': classrooms,
        'parents': parents,
        'page_title': f'O\'quvchini tahrirlash: {student.full_name}',
    }
    return render(request, 'students/student_form.html', context)


# ── Documents ─────────────────────────────────────────────────────────
@login_required
@students_required
def document_upload(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        doc_type_id = request.POST.get('doc_type')
        title = request.POST.get('title', '').strip()
        file = request.FILES.get('file')
        notes = request.POST.get('notes', '').strip()
        
        if file:
            doc = StudentDocument(
                student=student,
                doc_type_id=doc_type_id,
                title=title,
                file=file,
                uploaded_by=request.user,
                notes=notes
            )
            doc.save()
            messages.success(request, 'Hujjat yuklandi')
        else:
            messages.error(request, 'Fayl tanlanmadi')
    
    return redirect('students_detail', pk=pk)


@login_required
@students_required
def document_delete(request, doc_pk):
    doc = get_object_or_404(StudentDocument, pk=doc_pk)
    student_pk = doc.student.pk
    doc.delete()
    messages.success(request, 'Hujjat o\'chirildi')
    return redirect('students_detail', pk=student_pk)


@login_required
@students_required
def document_type_list(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        is_required = request.POST.get('is_required') == 'on'
        
        doc_type, created = DocumentType.objects.get_or_create(
            name=name,
            defaults={'is_required': is_required}
        )
        if not created:
            doc_type.is_required = is_required
            doc_type.save()
        
        messages.success(request, 'Hujjat turi saqlandi')
        return redirect('students_doc_types')
    
    doc_types = DocumentType.objects.all()
    context = {
        'doc_types': doc_types,
        'page_title': 'Hujjat turlari',
    }
    return render(request, 'students/document_type_list.html', context)


# ── Grades ────────────────────────────────────────────────────────────
@login_required
@students_required
def daily_grade_add(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        subject_id = request.POST.get('subject')
        grade = request.POST.get('grade')
        comment = request.POST.get('comment', '').strip()
        
        DailyGrade.objects.create(
            student_id=student_id,
            subject_id=subject_id,
            teacher=request.user,
            grade=grade,
            comment=comment
        )
        messages.success(request, 'Kunlik baholar qo\'shildi')
        return redirect('students_daily_grade')
    
    students = Student.objects.all()
    subjects = Subject.objects.all()
    context = {
        'students': students,
        'subjects': subjects,
        'page_title': 'Kunlik baholar qo\'shish',
    }
    return render(request, 'students/daily_grade_form.html', context)


@login_required
@students_required
def quarter_grade_add(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        subject_id = request.POST.get('subject')
        quarter_id = request.POST.get('quarter')
        grade = request.POST.get('grade')
        
        QuarterGrade.objects.update_or_create(
            student_id=student_id,
            subject_id=subject_id,
            quarter_id=quarter_id,
            defaults={'grade': grade, 'teacher': request.user}
        )
        messages.success(request, 'Chorak bahosi saqlandi')
        return redirect('students_quarter_grade')
    
    students = Student.objects.all()
    subjects = Subject.objects.all()
    quarters = Quarter.objects.select_related('academic_year').all()
    context = {
        'students': students,
        'subjects': subjects,
        'quarters': quarters,
        'page_title': 'Chorak baholar qo\'shish',
    }
    return render(request, 'students/quarter_grade_form.html', context)


@login_required
@students_required
def grade_results(request):
    student_id = request.GET.get('student')
    quarter_id = request.GET.get('quarter')
    
    students = Student.objects.all()
    quarters = Quarter.objects.select_related('academic_year').all()
    quarter_grades = QuarterGrade.objects.select_related('student', 'subject', 'quarter', 'teacher')
    
    if student_id:
        quarter_grades = quarter_grades.filter(student_id=student_id)
    if quarter_id:
        quarter_grades = quarter_grades.filter(quarter_id=quarter_id)
    
    # Calculate averages
    averages = quarter_grades.values('student__first_name', 'student__last_name').annotate(
        avg=Avg('grade')
    ).order_by('-avg')
    
    context = {
        'students': students,
        'quarters': quarters,
        'quarter_grades': quarter_grades,
        'averages': averages,
        'page_title': 'Baholar natijalari',
    }
    return render(request, 'students/grade_results.html', context)


# ── Attendance ────────────────────────────────────────────────────────
@login_required
@students_required
def attendance_mark(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        date_str = request.POST.get('date') or timezone.now().date().isoformat()
        status = request.POST.get('status', 'present')
        subject_id = request.POST.get('subject') or None
        note = request.POST.get('note', '').strip()
        
        Attendance.objects.update_or_create(
            student_id=student_id,
            date=date_str,
            subject_id=subject_id,
            defaults={'status': status, 'marked_by': request.user, 'note': note}
        )
        messages.success(request, 'Davomat saqlandi')
        return redirect('students_attendance')
    
    students = Student.objects.all()
    subjects = Subject.objects.all()
    context = {
        'students': students,
        'subjects': subjects,
        'page_title': 'Davomat belgilash',
    }
    return render(request, 'students/attendance_form.html', context)


# ── Homework ──────────────────────────────────────────────────────────
@login_required
@students_required
def homework_list(request):
    classroom_filter = request.GET.get('classroom')
    subject_filter = request.GET.get('subject')
    
    homeworks = Homework.objects.select_related('classroom', 'subject', 'teacher').all()
    
    if classroom_filter:
        homeworks = homeworks.filter(classroom_id=classroom_filter)
    if subject_filter:
        homeworks = homeworks.filter(subject_id=subject_filter)
    
    classrooms = Classroom.objects.all()
    subjects = Subject.objects.all()
    
    context = {
        'homeworks': homeworks,
        'classrooms': classrooms,
        'subjects': subjects,
        'page_title': 'Uy vazifalar',
    }
    return render(request, 'students/homework_list.html', context)


@login_required
@students_required
def homework_create(request):
    if request.method == 'POST':
        classroom_id = request.POST.get('classroom')
        subject_id = request.POST.get('subject')
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        due_date = request.POST.get('due_date')
        file = request.FILES.get('file')
        
        Homework.objects.create(
            classroom_id=classroom_id,
            subject_id=subject_id,
            teacher=request.user,
            title=title,
            description=description,
            due_date=due_date,
            file=file
        )
        messages.success(request, 'Uy vazifasi yaratildi')
        return redirect('students_homework_list')
    
    classrooms = Classroom.objects.all()
    subjects = Subject.objects.all()
    context = {
        'classrooms': classrooms,
        'subjects': subjects,
        'page_title': 'Yangi uy vazifasi',
    }
    return render(request, 'students/homework_form.html', context)


# ── Finance ───────────────────────────────────────────────────────────
@login_required
@students_required
def student_finance(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    contracts = student.contracts.select_related().all()
    payments = student.payments.select_related('contract', 'received_by').order_by('-payment_date')
    
    # Calculate balance
    total_paid = payments.aggregate(s=Sum('amount'))['s'] or 0
    total_fee = sum(c.effective_fee for c in contracts.filter(is_active=True))
    
    context = {
        'student': student,
        'contracts': contracts,
        'payments': payments,
        'total_paid': total_paid,
        'total_fee': total_fee,
        'balance': total_paid - total_fee,
        'page_title': f'O\'quvchi moliyasi: {student.full_name}',
    }
    return render(request, 'students/student_finance.html', context)


@login_required
@students_required
def payment_add(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        amount = request.POST.get('amount')
        method = request.POST.get('method', 'cash')
        payment_date = request.POST.get('payment_date') or timezone.now().date()
        month_for = request.POST.get('month_for') or payment_date
        contract_id = request.POST.get('contract') or None
        note = request.POST.get('note', '').strip()
        
        Payment.objects.create(
            student=student,
            contract_id=contract_id,
            amount=amount,
            method=method,
            payment_date=payment_date,
            month_for=month_for,
            received_by=request.user,
            note=note
        )
        messages.success(request, 'To\'lov qo\'shildi')
        return redirect('students_finance', pk=pk)
    
    contracts = student.contracts.filter(is_active=True)
    context = {
        'student': student,
        'contracts': contracts,
        'today': timezone.now().date(),
        'page_title': 'Yangi to\'lov',
    }
    return render(request, 'students/payment_form.html', context)


@login_required
@students_required
def contract_add(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        contract_number = request.POST.get('contract_number', '').strip()
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date') or None
        monthly_fee = request.POST.get('monthly_fee')
        discount_percent = request.POST.get('discount_percent', 0)
        discount_reason = request.POST.get('discount_reason', '').strip()
        notes = request.POST.get('notes', '').strip()
        
        Contract.objects.create(
            student=student,
            contract_number=contract_number,
            start_date=start_date,
            end_date=end_date,
            monthly_fee=monthly_fee,
            discount_percent=discount_percent,
            discount_reason=discount_reason,
            notes=notes
        )
        messages.success(request, 'Shartnoma yaratildi')
        return redirect('students_finance', pk=pk)
    
    context = {
        'student': student,
        'page_title': 'Yangi shartnoma',
    }
    return render(request, 'students/contract_form.html', context)


# ── SMS ───────────────────────────────────────────────────────────────
@login_required
@students_required
def sms_config(request):
    config, _ = SmsNotificationConfig.objects.get_or_create()
    
    if request.method == 'POST':
        config.day_of_month = request.POST.get('day_of_month', 5)
        config.is_active = request.POST.get('is_active') == 'on'
        config.message_template = request.POST.get('message_template', config.message_template)
        config.save()
        messages.success(request, 'SMS sozlamalari saqlandi')
        return redirect('students_sms_config')
    
    context = {
        'config': config,
        'page_title': 'SMS sozlamalari',
    }
    return render(request, 'students/sms_config.html', context)


@login_required
@students_required
def sms_send_now(request):
    if request.method == 'POST':
        config = SmsNotificationConfig.objects.first()
        if not config or not config.is_active:
            messages.error(request, 'SMS tizimi faollashtirilmagan')
            return redirect('students_sms_config')
        
        today = timezone.now().date()
        debtors = []
        
        for contract in Contract.objects.filter(is_active=True):
            paid = contract.payments.filter(
                payment_date__month=today.month,
                payment_date__year=today.year
            ).aggregate(s=Sum('amount'))['s'] or 0
            
            if paid < contract.effective_fee:
                debt = contract.effective_fee - paid
                for parent in contract.student.parents.all():
                    debtors.append({
                        'parent': parent,
                        'student': contract.student,
                        'contract': contract,
                        'debt': debt,
                        'phone': parent.phone
                    })
        
        # Send SMS (Twilio or local SMS gateway)
        sent_count = 0
        for d in debtors:
            message = config.message_template.format(
                parent_name=d['parent'].full_name,
                student_name=d['student'].full_name,
                contract_number=d['contract'].contract_number,
                debt_amount=d['debt']
            )
            
            # TODO: Real SMS sending via Twilio or local provider
            # For now, just log
            sms_log = SmsLog.objects.create(
                parent=d['parent'],
                student=d['student'],
                contract=d['contract'],
                phone=d['phone'],
                message=message,
                debt_amount=d['debt'],
                is_sent=False,
                error='SMS not sent (test mode)'
            )
        
            # Real SMS integration (Twilio example):
            # try:
            #     from twilio.rest import Client
            #     client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
            #     message_sent = client.messages.create(
            #         body=message,
            #         from_=settings.TWILIO_PHONE_NUMBER,
            #         to=d['phone']
            #     )
            #     sms_log.is_sent = True
            #     sms_log.save()
            #     sent_count += 1
            # except Exception as e:
            #     sms_log.error = str(e)
            #     sms_log.save()
        
        messages.success(request, f'{len(debtors)} ta qarzdor ota-onaga SMS yuborishga yuborildi')
        return redirect('students_sms_logs')
    
    return redirect('students_sms_config')


@login_required
@students_required
def sms_logs(request):
    logs = SmsLog.objects.select_related('parent', 'student', 'contract').order_by('-sent_at')
    
    paginator = Paginator(logs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'page_title': 'SMS xabarnomalar logi',
    }
    return render(request, 'students/sms_logs.html', context)


@csrf_exempt
@require_POST
def sms_daily_task(request):
    """Kunlik SMS yuborish task — faqat API token bilan"""
    api_token = request.headers.get('X-API-Token', '')
    expected_token = getattr(settings, 'SMS_TASK_API_TOKEN', '')
    if not expected_token or api_token != expected_token:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    config = SmsNotificationConfig.objects.first()
    if not config or not config.is_active:
        return JsonResponse({'error': 'SMS tizimi faol emas', 'sent': 0})
    
    today = timezone.now().date()
    
    # Faqat belgilangan kunda yuborish
    if today.day != config.day_of_month:
        return JsonResponse({'message': f"Bugun {today.day}-kuni, SMS {config.day_of_month}-kuni yuboriladi", 'sent': 0})
    
    debtors = []
    
    for contract in Contract.objects.filter(is_active=True):
        paid = contract.payments.filter(
            payment_date__month=today.month,
            payment_date__year=today.year
        ).aggregate(s=Sum('amount'))['s'] or 0
        
        if paid < contract.effective_fee:
            debt = contract.effective_fee - paid
            for parent in contract.student.parents.all():
                debtors.append({
                    'parent': parent,
                    'student': contract.student,
                    'contract': contract,
                    'debt': debt,
                    'phone': parent.phone
                })
    
    sent_count = 0
    for d in debtors:
        message = config.message_template.format(
            parent_name=d['parent'].full_name,
            student_name=d['student'].full_name,
            contract_number=d['contract'].contract_number,
            debt_amount=d['debt']
        )
        
        sms_log = SmsLog.objects.create(
            parent=d['parent'],
            student=d['student'],
            contract=d['contract'],
            phone=d['phone'],
            message=message,
            debt_amount=d['debt'],
            is_sent=False,
            error='Test mode'
        )
        
        # Real SMS integration
        # try:
        #     from twilio.rest import Client
        #     client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
        #     message_sent = client.messages.create(
        #         body=message,
        #         from_=settings.TWILIO_PHONE_NUMBER,
        #         to=d['phone']
        #     )
        #     sms_log.is_sent = True
        #     sms_log.save()
        #     sent_count += 1
        # except Exception as e:
        #     sms_log.error = str(e)
        #     sms_log.save()
    
    return JsonResponse({'message': f'{sent_count} ta SMS yuborildi', 'sent': sent_count, 'total_debtors': len(debtors)})


# ── Chat ──────────────────────────────────────────────────────────────
@login_required
@students_required
def chat_list(request):
    chat_groups = ChatGroup.objects.select_related('created_by', 'classroom').all()
    
    context = {
        'chat_groups': chat_groups,
        'page_title': 'Chat guruhlar',
    }
    return render(request, 'students/chat_list.html', context)


@login_required
@students_required
def chat_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        classroom_id = request.POST.get('classroom') or None
        student_ids = request.POST.getlist('students')
        member_ids = request.POST.getlist('members')
        
        chat = ChatGroup.objects.create(
            name=name,
            created_by=request.user,
            classroom_id=classroom_id
        )
        
        if student_ids:
            chat.students.add(*student_ids)
        if member_ids:
            chat.members.add(*member_ids)
        
        messages.success(request, 'Chat guruh yaratildi')
        return redirect('students_chat_detail', pk=chat.pk)
    
    classrooms = Classroom.objects.all()
    students = Student.objects.all()
    members = User.objects.filter(groups__name__in=['students_agent', 'students_manager', 'enrollment_agent', 'enrollment_manager']).distinct()
    
    context = {
        'classrooms': classrooms,
        'students': students,
        'members': members,
        'page_title': 'Yangi chat guruh',
    }
    return render(request, 'students/chat_form.html', context)


@login_required
@students_required
def chat_detail(request, pk):
    chat = get_object_or_404(
        ChatGroup.objects.select_related('created_by', 'classroom'),
        pk=pk
    )
    messages_list = chat.messages.select_related('sender').order_by('sent_at')
    
    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        file = request.FILES.get('file')
        
        if text or file:
            ChatMessage.objects.create(
                group=chat,
                sender=request.user,
                text=text,
                file=file
            )
            messages.success(request, 'Xabar yuborildi')
            return redirect('students_chat_detail', pk=pk)
    
    context = {
        'chat': chat,
        'messages': messages_list,
        'page_title': f'Chat: {chat.name}',
    }
    return render(request, 'students/chat_detail.html', context)


@login_required
@students_required
@require_POST
def chat_send(request, pk):
    chat = get_object_or_404(ChatGroup, pk=pk)
    text = request.POST.get('text', '').strip()
    file = request.FILES.get('file')
    
    if text or file:
        ChatMessage.objects.create(
            group=chat,
            sender=request.user,
            text=text,
            file=file
        )
        return JsonResponse({'status': 'ok'})
    
    return JsonResponse({'status': 'error', 'error': 'Empty message'}, status=400)


# ── Schedule ──────────────────────────────────────────────────────────
@login_required
@students_required
def schedule_list(request):
    classroom_filter = request.GET.get('classroom')
    
    schedules = Schedule.objects.select_related('classroom', 'subject', 'teacher').all()
    
    if classroom_filter:
        schedules = schedules.filter(classroom_id=classroom_filter)
    
    classrooms = Classroom.objects.all()
    
    context = {
        'schedules': schedules,
        'classrooms': classrooms,
        'page_title': 'Dars jadvali',
    }
    return render(request, 'students/schedule_list.html', context)


@login_required
@students_required
def schedule_create(request):
    if request.method == 'POST':
        classroom_id = request.POST.get('classroom')
        subject_id = request.POST.get('subject')
        teacher_id = request.POST.get('teacher') or None
        day_of_week = request.POST.get('day_of_week')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        
        Schedule.objects.create(
            classroom_id=classroom_id,
            subject_id=subject_id,
            teacher_id=teacher_id,
            day_of_week=day_of_week,
            start_time=start_time,
            end_time=end_time
        )
        messages.success(request, 'Dars jadvali qo\'shildi')
        return redirect('students_schedule')
    
    classrooms = Classroom.objects.all()
    subjects = Subject.objects.all()
    teachers = User.objects.filter(groups__name__in=['students_agent', 'students_manager', 'enrollment_agent', 'enrollment_manager']).distinct()
    
    context = {
        'classrooms': classrooms,
        'subjects': subjects,
        'teachers': teachers,
        'page_title': 'Yangi dars jadvali',
    }
    return render(request, 'students/schedule_form.html', context)


# ── Subjects ──────────────────────────────────────────────────────────
@login_required
@students_required
def subject_list(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        
        Subject.objects.get_or_create(name=name, defaults={'code': code})
        messages.success(request, 'Fan qo\'shildi')
        return redirect('students_subjects')
    
    subjects = Subject.objects.all()
    context = {
        'subjects': subjects,
        'page_title': 'Fanlar',
    }
    return render(request, 'students/subject_list.html', context)


# ── API endpoints ─────────────────────────────────────────────────────
@csrf_exempt
@require_POST
def api_add_student(request):
    """Meta Lead orqali o'quvchi yaratish uchun API"""
    try:
        data = json.loads(request.body)
        
        # Validate required fields
        if not data.get('phone'):
            return JsonResponse({'error': 'Telefon raqam kerak'}, status=400)
        
        # Check if parent exists
        parent, _ = Parent.objects.get_or_create(
            phone=data['phone'],
            defaults={
                'first_name': data.get('first_name', 'Noma\'lum'),
                'last_name': data.get('last_name', ''),
                'email': data.get('email', ''),
            }
        )
        
        # Create student
        student = Student.objects.create(
            first_name=data.get('child_first_name', 'Noma\'lum'),
            last_name=data.get('child_last_name', ''),
            birth_date=data.get('birth_date') or timezone.now().date(),
            gender=data.get('gender', 'M'),
            passport_or_id=data.get('passport_or_id', ''),
            erp_number=data.get('erp_number', ''),
            address=data.get('address', ''),
            medical_notes=data.get('medical_notes', ''),
        )
        student.parents.add(parent)
        
        # Link to lead if exists
        lead_id = data.get('lead_id')
        if lead_id:
            try:
                from enrollment.models import Lead
                lead = Lead.objects.get(pk=lead_id)
                lead.status = 'registered'
                lead.save()
            except Lead.DoesNotExist:
                pass
        
        return JsonResponse({
            'status': 'ok',
            'student_id': student.id,
            'student_name': student.full_name
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_POST
def api_send_sms(request):
    """SMS yuborish uchun API"""
    try:
        data = json.loads(request.body)
        phone = data.get('phone')
        message = data.get('message')
        
        if not phone or not message:
            return JsonResponse({'error': 'Telefon va xabar kerak'}, status=400)
        
        # TODO: Real SMS sending
        # For now, just log
        SmsLog.objects.create(
            phone=phone,
            message=message,
            is_sent=False,
            error='SMS not sent (test mode)'
        )
        
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ── Attendance Edit ───────────────────────────────────────────────────
@login_required
@students_required
def attendance_edit(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)

    if request.method == 'POST':
        attendance.status = request.POST.get('status', attendance.status)
        attendance.note = request.POST.get('note', '').strip()
        subject_id = request.POST.get('subject') or None
        attendance.subject_id = subject_id
        attendance.save()
        messages.success(request, 'Davomat yangilandi')
        return redirect('students_detail', pk=attendance.student.pk)

    subjects = Subject.objects.all()
    context = {
        'attendance': attendance,
        'subjects': subjects,
        'page_title': 'Davomatni tahrirlash',
    }
    return render(request, 'students/attendance_edit.html', context)


@login_required
@students_required
def attendance_delete(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    student_pk = attendance.student.pk
    attendance.delete()
    messages.success(request, 'Davomat o\'chirildi')
    return redirect('students_detail', pk=student_pk)


@login_required
@students_required
def attendance_list(request):
    """Barcha davomatlar ro'yxati — filter va export bilan"""
    classroom_filter = request.GET.get('classroom')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status', '')
    student_filter = request.GET.get('student', '')

    qs = Attendance.objects.select_related('student', 'subject', 'marked_by').all()

    if classroom_filter:
        qs = qs.filter(student__classroom_id=classroom_filter)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if student_filter:
        qs = qs.filter(student_id=student_filter)

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj': page_obj,
        'classrooms': Classroom.objects.all(),
        'students': Student.objects.all(),
        'classroom_filter': classroom_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_filter': status_filter,
        'student_filter': student_filter,
        'status_choices': Attendance.STATUS_CHOICES,
        'page_title': 'Davomat ro\'yxati',
    }
    return render(request, 'students/attendance_list.html', context)


# ── Excel Import / Export ─────────────────────────────────────────────
@login_required
@students_required
def export_students_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    qs = Student.objects.select_related('classroom').prefetch_related('parents').all()

    classroom_filter = request.GET.get('classroom')
    if classroom_filter:
        qs = qs.filter(classroom_id=classroom_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"

    headers = ['#', 'Familiya', 'Ism', 'Otasining ismi', 'Jinsi', 'Tug\'ilgan sana',
               'Sinf', 'Telefon (ota-ona)', 'Manzil', 'Faol', 'Ro\'yxatga olingan']
    hfill = PatternFill(start_color='1a73e8', end_color='1a73e8', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for i, s in enumerate(qs, 2):
        parent_phone = s.parents.first().phone if s.parents.exists() else ''
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=s.last_name)
        ws.cell(row=i, column=3, value=s.first_name)
        ws.cell(row=i, column=4, value=s.middle_name)
        ws.cell(row=i, column=5, value=s.get_gender_display())
        ws.cell(row=i, column=6, value=str(s.birth_date))
        ws.cell(row=i, column=7, value=str(s.classroom) if s.classroom else '')
        ws.cell(row=i, column=8, value=parent_phone)
        ws.cell(row=i, column=9, value=s.address)
        ws.cell(row=i, column=10, value='Ha' if s.is_active else "Yo'q")
        ws.cell(row=i, column=11, value=str(s.enrolled_date))

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
    return response


@login_required
@students_required
def import_students_excel(request):
    import openpyxl
    from django.utils.dateparse import parse_date

    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Fayl tanlanmadi')
            return redirect('students_list')

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            errors = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                try:
                    last_name = str(row[0] or '').strip()
                    first_name = str(row[1] or '').strip()
                    middle_name = str(row[2] or '').strip()
                    gender = 'M' if str(row[3] or 'M').upper() in ('M', 'ERKAK', "O'G'IL") else 'F'
                    birth_date = parse_date(str(row[4])) if row[4] else timezone.now().date()
                    classroom_name = str(row[5] or '').strip()
                    phone = str(row[6] or '').strip()

                    if not first_name or not last_name:
                        continue

                    classroom = None
                    if classroom_name:
                        classroom = Classroom.objects.filter(name=classroom_name).first()

                    student = Student.objects.create(
                        first_name=first_name,
                        last_name=last_name,
                        middle_name=middle_name,
                        gender=gender,
                        birth_date=birth_date or timezone.now().date(),
                        classroom=classroom,
                    )

                    if phone:
                        parent, _ = Parent.objects.get_or_create(
                            phone=phone,
                            defaults={'first_name': 'Ota-ona', 'last_name': last_name}
                        )
                        student.parents.add(parent)

                    created += 1
                except Exception as e:
                    errors.append(str(e))

            messages.success(request, f'{created} ta o\'quvchi import qilindi')
            if errors:
                messages.warning(request, f'{len(errors)} ta xato: {"; ".join(errors[:3])}')
        except Exception as e:
            messages.error(request, f'Fayl o\'qishda xato: {e}')

        return redirect('students_list')

    context = {'page_title': 'O\'quvchilarni import qilish'}
    return render(request, 'students/import_students.html', context)


@login_required
@students_required
def export_attendance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    classroom_filter = request.GET.get('classroom', '')

    qs = Attendance.objects.select_related('student', 'subject', 'marked_by').all()
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if classroom_filter:
        qs = qs.filter(student__classroom_id=classroom_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Davomat'

    headers = ['#', 'O\'quvchi', 'Sana', 'Fan', 'Holat', 'Izoh', 'Belgilagan']
    hfill = PatternFill(start_color='0f9d58', end_color='0f9d58', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    for i, a in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=str(a.student))
        ws.cell(row=i, column=3, value=str(a.date))
        ws.cell(row=i, column=4, value=str(a.subject) if a.subject else '')
        ws.cell(row=i, column=5, value=a.get_status_display())
        ws.cell(row=i, column=6, value=a.note)
        ws.cell(row=i, column=7, value=a.marked_by.username if a.marked_by else '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'
    return response


@login_required
@students_required
def export_payments_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    qs = Payment.objects.select_related('student', 'contract', 'received_by').all()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        qs = qs.filter(payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(payment_date__lte=date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "To'lovlar"

    headers = ['#', 'O\'quvchi', 'Shartnoma', 'Summa', 'Usul', 'Sana', 'Oy uchun', 'Tranzaksiya ID', 'Izoh', 'Qabul qildi']
    hfill = PatternFill(start_color='ea4335', end_color='ea4335', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=str(p.student))
        ws.cell(row=i, column=3, value=p.contract.contract_number if p.contract else '')
        ws.cell(row=i, column=4, value=float(p.amount))
        ws.cell(row=i, column=5, value=p.get_method_display())
        ws.cell(row=i, column=6, value=str(p.payment_date))
        ws.cell(row=i, column=7, value=str(p.month_for))
        ws.cell(row=i, column=8, value=p.transaction_id)
        ws.cell(row=i, column=9, value=p.note)
        ws.cell(row=i, column=10, value=p.received_by.username if p.received_by else '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
    return response


# ── Online Payment Gateway Views ──────────────────────────────────────

@login_required
@students_required
def online_payment_list(request):
    qs = OnlinePayment.objects.select_related('student', 'contract').all()
    provider_filter = request.GET.get('provider', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if provider_filter:
        qs = qs.filter(provider=provider_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    paginator = Paginator(qs, 30)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj': page_obj,
        'provider_filter': provider_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'providers': OnlinePayment.PROVIDER_CHOICES,
        'statuses': OnlinePayment.STATUS_CHOICES,
        'page_title': 'Online to\'lovlar',
    }
    return render(request, 'students/online_payment_list.html', context)


# ── Payme Webhook ─────────────────────────────────────────────────────
@csrf_exempt
def payme_webhook(request):
    """Payme (merchant API) webhook handler"""
    import json as _json
    try:
        data = _json.loads(request.body)
        method = data.get('method', '')
        params = data.get('params', {})
        _id = data.get('id', 1)

        if method == 'CheckPerformTransaction':
            order_id = params.get('account', {}).get('order_id', '')
            contract = Contract.objects.filter(contract_number=order_id, is_active=True).first()
            if not contract:
                return JsonResponse({'id': _id, 'error': {'code': -31050, 'message': 'Order not found'}})
            return JsonResponse({'id': _id, 'result': {'allow': True}})

        elif method == 'CreateTransaction':
            order_id = params.get('account', {}).get('order_id', '')
            transaction_id = params.get('id', '')
            amount = params.get('amount', 0) / 100  # tiyin -> so'm

            contract = Contract.objects.filter(contract_number=order_id, is_active=True).first()
            if not contract:
                return JsonResponse({'id': _id, 'error': {'code': -31050, 'message': 'Order not found'}})

            op, created = OnlinePayment.objects.get_or_create(
                transaction_id=transaction_id,
                defaults={
                    'provider': 'payme',
                    'order_id': order_id,
                    'student': contract.student,
                    'contract': contract,
                    'amount': amount,
                    'status': 'pending',
                    'raw_request': data,
                }
            )
            return JsonResponse({'id': _id, 'result': {
                'create_time': int(op.created_at.timestamp() * 1000),
                'transaction': str(op.pk),
                'state': 1,
            }})

        elif method == 'PerformTransaction':
            transaction_id = params.get('id', '')
            op = OnlinePayment.objects.filter(transaction_id=transaction_id, provider='payme').first()
            if not op:
                return JsonResponse({'id': _id, 'error': {'code': -31003, 'message': 'Transaction not found'}})

            if op.status != 'paid':
                op.status = 'paid'
                op.confirmed_at = timezone.now()
                op.save()
                _confirm_online_payment(op)

            return JsonResponse({'id': _id, 'result': {
                'transaction': str(op.pk),
                'perform_time': int(op.confirmed_at.timestamp() * 1000),
                'state': 2,
            }})

        elif method == 'CancelTransaction':
            transaction_id = params.get('id', '')
            op = OnlinePayment.objects.filter(transaction_id=transaction_id).first()
            if op:
                op.status = 'cancelled'
                op.save()
            return JsonResponse({'id': _id, 'result': {'state': -1, 'cancel_time': int(timezone.now().timestamp() * 1000), 'transaction': str(op.pk) if op else '0'}})

        elif method == 'CheckTransaction':
            transaction_id = params.get('id', '')
            op = OnlinePayment.objects.filter(transaction_id=transaction_id).first()
            if not op:
                return JsonResponse({'id': _id, 'error': {'code': -31003, 'message': 'Transaction not found'}})
            state_map = {'pending': 1, 'paid': 2, 'cancelled': -1, 'failed': -2}
            return JsonResponse({'id': _id, 'result': {
                'create_time': int(op.created_at.timestamp() * 1000),
                'perform_time': int(op.confirmed_at.timestamp() * 1000) if op.confirmed_at else 0,
                'cancel_time': 0,
                'transaction': str(op.pk),
                'state': state_map.get(op.status, 1),
                'reason': None,
            }})

    except Exception as e:
        return JsonResponse({'id': 1, 'error': {'code': -32400, 'message': str(e)}})

    return JsonResponse({'id': 1, 'error': {'code': -32601, 'message': 'Method not found'}})


# ── Click Webhook ─────────────────────────────────────────────────────
@csrf_exempt
def click_webhook(request):
    """Click (SHOP API) webhook handler"""
    if request.method != 'POST':
        return JsonResponse({'error': -1, 'error_note': 'Method not allowed'})

    try:
        action = int(request.POST.get('action', -1))
        click_trans_id = request.POST.get('click_trans_id', '')
        merchant_trans_id = request.POST.get('merchant_trans_id', '')  # order_id
        amount = float(request.POST.get('amount', 0))
        error = int(request.POST.get('error', 0))

        contract = Contract.objects.filter(contract_number=merchant_trans_id, is_active=True).first()
        if not contract:
            return JsonResponse({'error': -5, 'error_note': 'User does not exist'})

        if action == 0:  # Prepare
            op, _ = OnlinePayment.objects.get_or_create(
                transaction_id=click_trans_id,
                defaults={
                    'provider': 'click',
                    'order_id': merchant_trans_id,
                    'student': contract.student,
                    'contract': contract,
                    'amount': amount,
                    'status': 'pending',
                    'raw_request': dict(request.POST),
                }
            )
            return JsonResponse({
                'click_trans_id': click_trans_id,
                'merchant_trans_id': merchant_trans_id,
                'merchant_prepare_id': op.pk,
                'error': 0,
                'error_note': 'Success',
            })

        elif action == 1:  # Complete
            op = OnlinePayment.objects.filter(transaction_id=click_trans_id, provider='click').first()
            if not op:
                return JsonResponse({'error': -6, 'error_note': 'Transaction not found'})

            if error == 0 and op.status != 'paid':
                op.status = 'paid'
                op.confirmed_at = timezone.now()
                op.save()
                _confirm_online_payment(op)
            elif error < 0:
                op.status = 'cancelled'
                op.save()

            return JsonResponse({
                'click_trans_id': click_trans_id,
                'merchant_trans_id': merchant_trans_id,
                'merchant_confirm_id': op.pk,
                'error': 0,
                'error_note': 'Success',
            })

    except Exception as e:
        return JsonResponse({'error': -9, 'error_note': str(e)})

    return JsonResponse({'error': -1, 'error_note': 'Unknown action'})


# ── Uzum (Apelsin) Webhook ────────────────────────────────────────────
@csrf_exempt
def uzum_webhook(request):
    """Uzum Bank webhook handler"""
    import json as _json
    try:
        data = _json.loads(request.body)
        service_id = data.get('serviceId', '')
        transaction_id = str(data.get('transactionId', ''))
        order_id = str(data.get('params', {}).get('account', ''))
        amount = float(data.get('amount', 0)) / 100
        status = data.get('status', '')

        contract = Contract.objects.filter(contract_number=order_id, is_active=True).first()
        if not contract:
            return JsonResponse({'status': -1, 'desc': 'Order not found'})

        op, _ = OnlinePayment.objects.get_or_create(
            transaction_id=transaction_id,
            defaults={
                'provider': 'uzum',
                'order_id': order_id,
                'student': contract.student,
                'contract': contract,
                'amount': amount,
                'status': 'pending',
                'raw_request': data,
            }
        )

        if status == 'CONFIRMED' and op.status != 'paid':
            op.status = 'paid'
            op.confirmed_at = timezone.now()
            op.save()
            _confirm_online_payment(op)

        return JsonResponse({'status': 0, 'desc': 'Success'})
    except Exception as e:
        return JsonResponse({'status': -1, 'desc': str(e)})


def _confirm_online_payment(op: OnlinePayment):
    """
    Online to'lov tasdiqlanganda:
    1. Payment yaratish
    2. Buxgalteriyaga Transaction yozish
    """
    from accounting.models import Transaction, CashRegister

    if op.payment:
        return  # Allaqachon yaratilgan

    month_for = op.month_for or timezone.now().date().replace(day=1)

    payment = Payment.objects.create(
        student=op.student,
        contract=op.contract,
        amount=op.amount,
        method=op.provider,
        payment_date=op.confirmed_at.date() if op.confirmed_at else timezone.now().date(),
        month_for=month_for,
        note=f'{op.get_provider_display()} orqali online to\'lov. TxID: {op.transaction_id}',
        transaction_id=op.transaction_id,
        is_confirmed=True,
    )
    op.payment = payment
    op.save(update_fields=['payment'])

    # Buxgalteriyaga yozish
    try:
        cash_register = CashRegister.objects.filter(
            name__icontains=op.provider
        ).first() or CashRegister.objects.first()

        if cash_register:
            Transaction.objects.create(
                cash_register=cash_register,
                amount=op.amount,
                transaction_type='income',
                description=(
                    f"{op.get_provider_display()} to'lov | "
                    f"O'quvchi: {op.student} | "
                    f"Shartnoma: {op.contract.contract_number if op.contract else '—'} | "
                    f"TxID: {op.transaction_id}"
                ),
            )
            cash_register.balance += op.amount
            cash_register.save(update_fields=['balance'])
    except Exception:
        pass  # Buxgalteriya xatosi to'lovni bloklamasin


# ── Payment Init (redirect to gateway) ───────────────────────────────
@login_required
@students_required
def payment_init(request, pk):
    """To'lov tizimiga yo'naltirish sahifasi"""
    student = get_object_or_404(Student, pk=pk)
    contracts = student.contracts.filter(is_active=True)

    context = {
        'student': student,
        'contracts': contracts,
        'page_title': f'Online to\'lov: {student.full_name}',
        'providers': OnlinePayment.PROVIDER_CHOICES,
    }
    return render(request, 'students/payment_init.html', context)


# ══════════════════════════════════════════════════════════════════════
# SINF JURNALI VA DAVOMAT MODULI
# ══════════════════════════════════════════════════════════════════════

@login_required
@students_required
def classroom_journal(request, pk):
    """Sinf jurnali — joriy dars, o'quvchilar ro'yxati, davomat"""
    from datetime import datetime as dt
    classroom = get_object_or_404(
        Classroom.objects.select_related('grade', 'academic_year', 'homeroom_teacher'),
        pk=pk
    )
    students = classroom.students.filter(is_active=True).order_by('last_name', 'first_name')
    today = timezone.now().date()
    now_time = timezone.now().time()
    today_weekday = today.isoweekday()  # 1=Mon..6=Sat

    # Bugungi dars jadvali
    today_schedules = classroom.schedules.filter(
        day_of_week=today_weekday
    ).select_related('subject', 'teacher', 'period').order_by(
        'period__lesson_number', 'start_time'
    )

    # Joriy dars — hozirgi vaqtga mos keluvchi
    current_schedule = None
    next_schedule = None
    for sch in today_schedules:
        s = sch.effective_start
        e = sch.effective_end
        if s and e:
            if s <= now_time <= e:
                current_schedule = sch
                break
            elif s > now_time and next_schedule is None:
                next_schedule = sch

    # Tanlangan dars (URL param yoki joriy)
    selected_schedule_id = request.GET.get('schedule')
    selected_schedule = None
    if selected_schedule_id:
        selected_schedule = today_schedules.filter(pk=selected_schedule_id).first()
    if not selected_schedule:
        selected_schedule = current_schedule or (today_schedules.first() if today_schedules.exists() else None)

    # Mavjud davomat yozuvlari (bugun, tanlangan fan uchun)
    existing_attendance = {}
    if selected_schedule:
        for a in Attendance.objects.filter(
            student__classroom=classroom,
            date=today,
            subject=selected_schedule.subject
        ):
            existing_attendance[a.student_id] = a

    # POST — davomat saqlash
    if request.method == 'POST':
        schedule_id = request.POST.get('schedule_id')
        sch = get_object_or_404(Schedule, pk=schedule_id, classroom=classroom)
        for student in students:
            status = request.POST.get(f'status_{student.pk}', 'present')
            note = request.POST.get(f'note_{student.pk}', '').strip()
            Attendance.objects.update_or_create(
                student=student,
                date=today,
                subject=sch.subject,
                defaults={'status': status, 'marked_by': request.user, 'note': note}
            )
        messages.success(request, f'Davomat saqlandi: {sch.subject.name}')
        return redirect(f"{request.path}?schedule={schedule_id}")

    # Haftalik jadval (barcha kunlar)
    week_schedule = {}
    for day_num, day_name in Schedule.DAYS:
        week_schedule[day_name] = classroom.schedules.filter(
            day_of_week=day_num
        ).select_related('subject', 'teacher', 'period').order_by(
            'period__lesson_number', 'start_time'
        )

    # O'quvchilar uchun bugungi davomat statistikasi
    today_stats = {
        'present': Attendance.objects.filter(student__classroom=classroom, date=today, status='present').values('student').distinct().count(),
        'absent': Attendance.objects.filter(student__classroom=classroom, date=today, status='absent').values('student').distinct().count(),
        'late': Attendance.objects.filter(student__classroom=classroom, date=today, status='late').values('student').distinct().count(),
    }

    context = {
        'classroom': classroom,
        'students': students,
        'today': today,
        'today_schedules': today_schedules,
        'current_schedule': current_schedule,
        'next_schedule': next_schedule,
        'selected_schedule': selected_schedule,
        'existing_attendance': existing_attendance,
        'week_schedule': week_schedule,
        'today_stats': today_stats,
        'page_title': f'Sinf jurnali: {classroom.name}',
    }
    return render(request, 'students/classroom_journal.html', context)


@login_required
@students_required
def attendance_by_date(request, pk):
    """Sinf uchun sana bo'yicha davomat ko'rish"""
    classroom = get_object_or_404(Classroom, pk=pk)
    date_str = request.GET.get('date', timezone.now().date().isoformat())
    try:
        from datetime import date as date_type
        sel_date = date_type.fromisoformat(date_str)
    except ValueError:
        sel_date = timezone.now().date()

    students = classroom.students.filter(is_active=True).order_by('last_name', 'first_name')
    attendances = Attendance.objects.filter(
        student__classroom=classroom, date=sel_date
    ).select_related('student', 'subject')

    att_map = {}
    for a in attendances:
        att_map.setdefault(a.student_id, []).append(a)

    context = {
        'classroom': classroom,
        'students': students,
        'sel_date': sel_date,
        'att_map': att_map,
        'page_title': f'Davomat: {classroom.name} — {sel_date}',
    }
    return render(request, 'students/attendance_by_date.html', context)


# ══════════════════════════════════════════════════════════════════════
# DARS JADVALI CRUD
# ══════════════════════════════════════════════════════════════════════

@login_required
@students_required
def schedule_edit(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    if request.method == 'POST':
        schedule.subject_id = request.POST.get('subject')
        schedule.teacher_id = request.POST.get('teacher') or None
        schedule.day_of_week = request.POST.get('day_of_week')
        schedule.period_id = request.POST.get('period') or None
        schedule.start_time = request.POST.get('start_time') or None
        schedule.end_time = request.POST.get('end_time') or None
        schedule.room = request.POST.get('room', '').strip()
        schedule.save()
        messages.success(request, 'Dars jadvali yangilandi')
        return redirect('students_classroom_journal', pk=schedule.classroom.pk)

    subjects = Subject.objects.all()
    teachers = User.objects.filter(
        groups__name__in=['students_agent', 'students_manager', 'enrollment_agent', 'enrollment_manager']
    ).distinct()
    periods = LessonPeriod.objects.all()
    context = {
        'schedule': schedule,
        'subjects': subjects,
        'teachers': teachers,
        'periods': periods,
        'days': Schedule.DAYS,
        'page_title': 'Dars jadvalini tahrirlash',
    }
    return render(request, 'students/schedule_form_full.html', context)


@login_required
@students_required
def schedule_delete(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    classroom_pk = schedule.classroom.pk
    schedule.delete()
    messages.success(request, "Dars o'chirildi")
    return redirect('students_classroom_journal', pk=classroom_pk)


@login_required
@students_required
def schedule_create_for_classroom(request, classroom_pk):
    classroom = get_object_or_404(Classroom, pk=classroom_pk)
    if request.method == 'POST':
        Schedule.objects.create(
            classroom=classroom,
            subject_id=request.POST.get('subject'),
            teacher_id=request.POST.get('teacher') or None,
            day_of_week=request.POST.get('day_of_week'),
            period_id=request.POST.get('period') or None,
            start_time=request.POST.get('start_time') or None,
            end_time=request.POST.get('end_time') or None,
            room=request.POST.get('room', '').strip(),
        )
        messages.success(request, 'Dars qo\'shildi')
        return redirect('students_classroom_journal', pk=classroom_pk)

    subjects = Subject.objects.all()
    teachers = User.objects.filter(
        groups__name__in=['students_agent', 'students_manager', 'enrollment_agent', 'enrollment_manager']
    ).distinct()
    periods = LessonPeriod.objects.all()
    context = {
        'classroom': classroom,
        'subjects': subjects,
        'teachers': teachers,
        'periods': periods,
        'days': Schedule.DAYS,
        'page_title': f'{classroom.name} — Yangi dars qo\'shish',
    }
    return render(request, 'students/schedule_form_full.html', context)


# ══════════════════════════════════════════════════════════════════════
# DARS SOATLARI (LessonPeriod) CRUD
# ══════════════════════════════════════════════════════════════════════

@login_required
@students_required
def lesson_period_list(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            LessonPeriod.objects.create(
                level=request.POST.get('level'),
                lesson_number=request.POST.get('lesson_number'),
                start_time=request.POST.get('start_time'),
                end_time=request.POST.get('end_time'),
                label=request.POST.get('label', '').strip(),
            )
            messages.success(request, 'Dars soati qo\'shildi')
        elif action == 'delete':
            LessonPeriod.objects.filter(pk=request.POST.get('pk')).delete()
            messages.success(request, 'O\'chirildi')
        return redirect('students_lesson_periods')

    periods = LessonPeriod.objects.all()
    context = {
        'periods': periods,
        'levels': LessonPeriod.LEVEL_CHOICES,
        'page_title': 'Dars soatlari',
    }
    return render(request, 'students/lesson_period_list.html', context)


@login_required
@students_required
def lesson_period_edit(request, pk):
    period = get_object_or_404(LessonPeriod, pk=pk)
    if request.method == 'POST':
        period.level = request.POST.get('level')
        period.lesson_number = request.POST.get('lesson_number')
        period.start_time = request.POST.get('start_time')
        period.end_time = request.POST.get('end_time')
        period.label = request.POST.get('label', '').strip()
        period.save()
        messages.success(request, 'Dars soati yangilandi')
        return redirect('students_lesson_periods')

    context = {
        'period': period,
        'levels': LessonPeriod.LEVEL_CHOICES,
        'page_title': 'Dars soatini tahrirlash',
    }
    return render(request, 'students/lesson_period_edit.html', context)
