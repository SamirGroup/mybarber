import io
import json
from decimal import Decimal, InvalidOperation
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, Avg
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    FinishedGoodsInventory, Product, ProductCategory, ProductDayBalance,
    ProductionLog, RawMaterial, Recipe, RecipeItem,
)


def _opening_qty_from_previous_day(prev_date, product_id, physical_fallback):
    prev = ProductDayBalance.objects.filter(balance_date=prev_date, product_id=product_id).first()
    if prev and prev.closing_qty is not None:
        return prev.closing_qty
    return physical_fallback


def _run_close_inventory_day(d):
    """Kunning yakuni: har mahsulot uchun closing_qty = joriy fizik qoldiq."""
    prev = d - timedelta(days=1)
    with transaction.atomic():
        for inv in FinishedGoodsInventory.objects.select_related('product'):
            opening_default = _opening_qty_from_previous_day(prev, inv.product_id, inv.stock)
            row, _ = ProductDayBalance.objects.get_or_create(
                balance_date=d,
                product_id=inv.product_id,
                defaults={'opening_qty': opening_default},
            )
            row.closing_qty = inv.stock
            row.closed_at = timezone.now()
            row.save(update_fields=['closing_qty', 'closed_at'])


def _run_carry_inventory_forward(from_date, to_date):
    """Keyingi kun kirishi: opening_qty = oldingi kun closing_qty (carried-forward)."""
    with transaction.atomic():
        for inv in FinishedGoodsInventory.objects.select_related('product'):
            prev_row = ProductDayBalance.objects.filter(balance_date=from_date, product_id=inv.product_id).first()
            opening = prev_row.closing_qty if prev_row and prev_row.closing_qty is not None else inv.stock
            ProductDayBalance.objects.update_or_create(
                balance_date=to_date,
                product_id=inv.product_id,
                defaults={'opening_qty': opening},
            )


def _can_access(user, *roles):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=roles).exists()


def _date_range(request):
    """Return (date_from, date_to) from GET params or sensible defaults."""
    period = request.GET.get('period', 'today')
    today = timezone.localdate()
    if period == 'week':
        return today - timedelta(days=6), today
    if period == 'custom':
        try:
            df = date.fromisoformat(request.GET.get('date_from', ''))
            dt = date.fromisoformat(request.GET.get('date_to', ''))
            return df, dt
        except ValueError:
            pass
    return today, today


@login_required
def production_dashboard(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')

    raw_materials = RawMaterial.objects.all().order_by('name')
    finished_goods = FinishedGoodsInventory.objects.select_related('product').order_by('product__name')
    products = Product.objects.select_related('category', 'recipe').all().order_by('name')

    # ── POST: handle actions ──────────────────────────────────────────
    if request.method == 'POST':
        if 'add_production' in request.POST:
            product_id = request.POST.get('product')
            batches_raw = request.POST.get('batches', '1')
            baker_name = request.POST.get('baker_name', '').strip() or 'System'
            timer_min = int(request.POST.get('timer_minutes', 0) or 0)

            try:
                batches = int(batches_raw)
                if batches <= 0:
                    raise ValueError("Partiya soni 0 dan katta bo'lishi kerak.")

                product = Product.objects.get(id=product_id)
                recipe = product.recipe
                quantity = recipe.batch_size * batches
                batch_multiplier = Decimal(str(batches))

                with transaction.atomic():
                    for item in recipe.items.all():
                        # quantity_grams saqlangan bo'lsa, uni asosiy birlikka (kg/liter) o'girib ishlatamiz
                        if item.quantity_grams is not None:
                            mat_unit = item.raw_material.unit.lower()
                            if mat_unit == 'kg':
                                required = (item.quantity_grams / Decimal('1000')) * batch_multiplier
                            elif mat_unit == 'liter':
                                required = (item.quantity_grams / Decimal('1000')) * batch_multiplier
                            else:
                                required = item.quantity * batch_multiplier
                        else:
                            required = item.quantity * batch_multiplier

                        if item.raw_material.stock < required:
                            raise ValueError(
                                f"{item.raw_material.name} yetarli emas "
                                f"(kerak: {required:.3f} {item.raw_material.unit}, "
                                f"bor: {item.raw_material.stock} {item.raw_material.unit})."
                            )
                        item.raw_material.stock -= required
                        item.raw_material.save()

                    immediate = timer_min == 0
                    log = ProductionLog.objects.create(
                        product=product,
                        quantity=quantity,
                        batches=batches,
                        baker_name=baker_name,
                        timer_minutes=timer_min,
                        timer_started_at=timezone.now(),
                        is_done=immediate,
                    )
                    if immediate:
                        fg, _ = FinishedGoodsInventory.objects.get_or_create(product=product)
                        fg.stock += quantity
                        fg.produced_at = timezone.now()
                        fg.save()

                if immediate:
                    messages.success(request, f"{quantity} dona {product.name} omborga qo'shildi (darhol tayyor).")
                else:
                    messages.success(request, f"{quantity} dona {product.name} ishlab chiqarish jarayoniga qo'shildi ({timer_min} min).")
            except Product.DoesNotExist:
                messages.error(request, "Mahsulot topilmadi.")
            except Recipe.DoesNotExist:
                messages.error(request, "Retsept mavjud emas.")
            except (ValueError, InvalidOperation) as e:
                messages.error(request, str(e))
            
            pq = request.POST.get('preserve_query', '').strip()
            if pq:
                return redirect(f'/production/?{pq}')
            return redirect('production_dashboard')

        elif 'edit_raw_material' in request.POST:
            mat_id = request.POST.get('mat_id')
            stock = request.POST.get('stock')
            try:
                mat = RawMaterial.objects.get(id=mat_id)
                mat.stock = Decimal(stock)
                mat.save()
                messages.success(request, f"'{mat.name}' qoldig'i tahrirlandi.")
            except Exception as e:
                messages.error(request, str(e))
            return redirect('production_dashboard')

        elif 'edit_finished_good' in request.POST:
            fg_id = request.POST.get('fg_id')
            stock = request.POST.get('stock')
            try:
                fg = FinishedGoodsInventory.objects.get(id=fg_id)
                fg.stock = int(stock)
                fg.save()
                messages.success(request, f"'{fg.product.name}' tayyor qoldig'i tahrirlandi.")
            except Exception as e:
                messages.error(request, str(e))
            return redirect('production_dashboard')

    # ── Date filter for logs ──────────────────────────────────────────
    date_from, date_to = _date_range(request)
    recent_logs = (
        ProductionLog.objects
        .select_related('product')
        .filter(date__date__gte=date_from, date__date__lte=date_to)
        .order_by('-date')[:50]
    )

    # ── Excel export ──────────────────────────────────────────────────
    if request.GET.get('export') == 'excel':
        return _export_production_excel(date_from, date_to)

    today = timezone.localdate()
    # Count only completed production logs in KPI totals.
    # In-progress timer logs must not affect "Umumiy ishlab chiqarish" until marked done.
    today_total = (
        ProductionLog.objects.filter(date__date=today, is_done=True).aggregate(t=Sum('quantity'))['t'] or 0
    )
    total_finished = finished_goods.aggregate(t=Sum('stock'))['t'] or 0
    low_material_count = raw_materials.filter(stock__lt=10).count()
    recipes_count = Recipe.objects.count()
    products_count = products.count()
    recipe_coverage = round((recipes_count / products_count) * 100, 1) if products_count else 0

    finished_stock_value = sum(
        (fg.product.price * fg.stock) for fg in finished_goods
    ) or Decimal('0')

    produced_done_rows = (
        ProductionLog.objects.filter(
            is_done=True,
            date__date__gte=date_from,
            date__date__lte=date_to,
        )
        .values('product_id')
        .annotate(total_done=Sum('quantity'))
    )
    produced_done_by_product = {
        row['product_id']: int(row['total_done'] or 0)
        for row in produced_done_rows
    }

    finished_rows = [
        {
            'item': fg,
            'produced_done': produced_done_by_product.get(fg.product_id, 0),
            'line_total': fg.product.price * fg.stock,
        }
        for fg in finished_goods
    ]
    critical_materials = raw_materials.filter(stock__lt=10).order_by('stock')[:8]

    # Production mix chart should also reflect only completed output.
    product_mix = (
        ProductionLog.objects.filter(is_done=True)
        .values('product__name')
        .annotate(total=Sum('quantity'))
        .order_by('-total')[:8]
    )
    mix_labels = [r['product__name'] for r in product_mix]
    mix_data = [r['total'] for r in product_mix]

    # Active timers
    active_logs = ProductionLog.objects.filter(is_done=False, timer_minutes__gt=0).select_related('product').order_by('-date')

    active_logs_payload = [
        {
            'id': log.id,
            'product': log.product.name,
            'quantity': log.quantity,
            'batches': log.batches,
            'timer_minutes': log.timer_minutes,
            'started_iso': log.timer_started_at.isoformat() if log.timer_started_at else None,
        }
        for log in active_logs
    ]

    recipe_batches_payload = []
    for p in products:
        batch_sz = None
        try:
            batch_sz = p.recipe.batch_size
        except Recipe.DoesNotExist:
            pass
        recipe_batches_payload.append({'id': p.id, 'batch_size': batch_sz})

    context = {
        'raw_materials': raw_materials,
        'finished_goods': finished_goods,
        'recent_logs': recent_logs,
        'products': products,
        'today_total_production': today_total,
        'total_finished_stock': total_finished,
        'low_material_count': low_material_count,
        'products_without_recipe': max(products_count - recipes_count, 0),
        'finished_stock_value': finished_stock_value,
        'finished_rows': finished_rows,
        'critical_materials': critical_materials,
        'mix_labels': mix_labels,
        'mix_data': mix_data,
        'recipe_coverage': recipe_coverage,
        'active_logs': active_logs,
        'active_logs_payload': active_logs_payload,
        'recipe_batches_payload': recipe_batches_payload,
        'date_from': date_from,
        'date_to': date_to,
        'period': request.GET.get('period', 'today'),
    }
    return render(request, 'production.html', context)


def _export_production_excel(date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan. pip install openpyxl", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ishlab chiqarish"

    header_fill = PatternFill("solid", fgColor="D4A373")
    bold = Font(bold=True)
    headers = ['Sana', 'Mahsulot', 'Miqdor (dona)', 'Partiya', 'Novvoy', 'Timer (min)', 'Holat']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    logs = ProductionLog.objects.filter(
        date__date__gte=date_from, date__date__lte=date_to
    ).select_related('product').order_by('-date')

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.date.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row_idx, column=2, value=log.product.name)
        ws.cell(row=row_idx, column=3, value=log.quantity)
        ws.cell(row=row_idx, column=4, value=log.batches)
        ws.cell(row=row_idx, column=5, value=log.baker_name or '')
        ws.cell(row=row_idx, column=6, value=log.timer_minutes)
        ws.cell(row=row_idx, column=7, value='Tayyor' if log.is_done else 'Jarayonda')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="production_{date_from}_{date_to}.xlsx"'
    return resp


@login_required
def mark_production_done(request, log_id):
    """AJAX: mark a production log as done and add to inventory."""
    if not _can_access(request.user, 'seller', 'production_manager'):
        return JsonResponse({'ok': False}, status=403)
    try:
        log = ProductionLog.objects.select_related('product').get(id=log_id, is_done=False)
        with transaction.atomic():
            log.is_done = True
            log.save()
            fg, _ = FinishedGoodsInventory.objects.get_or_create(product=log.product)
            fg.stock += log.quantity
            fg.produced_at = timezone.now()
            fg.save()
        return JsonResponse({'ok': True, 'product': log.product.name, 'qty': log.quantity})
    except ProductionLog.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Topilmadi'}, status=404)


@login_required
def manage_products(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')

    products = Product.objects.select_related('category', 'inventory').order_by('name')
    raw_materials = RawMaterial.objects.all().order_by('name')
    recipes = Recipe.objects.select_related('product').prefetch_related('items__raw_material').all()
    categories = ProductCategory.objects.all().order_by('name')

    if request.method == 'POST':
        action = request.POST.get('action')
        manage_redirect_balance_date = None

        if action == 'add_category':
            name = request.POST.get('cat_name', '').strip()
            if name:
                ProductCategory.objects.get_or_create(name=name)
                messages.success(request, f"Kategoriya '{name}' qo'shildi.")
            else:
                messages.error(request, "Kategoriya nomi majburiy.")

        elif action == 'edit_category':
            cid = request.POST.get('category_id')
            try:
                cat = ProductCategory.objects.get(id=cid)
                new_name = request.POST.get('cat_name', '').strip()
                if new_name:
                    cat.name = new_name
                    cat.save()
                    messages.success(request, f"Kategoriya '{cat.name}' yangilandi.")
                else:
                    messages.error(request, "Kategoriya nomi majburiy.")
            except ProductCategory.DoesNotExist:
                messages.error(request, "Kategoriya topilmadi.")

        elif action == 'delete_category':
            cid = request.POST.get('category_id')
            try:
                ProductCategory.objects.get(id=cid).delete()
                messages.success(request, "Kategoriya o'chirildi.")
            except ProductCategory.DoesNotExist:
                messages.error(request, "Kategoriya topilmadi.")

        elif action == 'add_product':
            name = request.POST.get('name', '').strip()
            price_raw = request.POST.get('price', '').strip()
            cat_id = request.POST.get('category', '')
            image = request.FILES.get('image')
            try:
                price = Decimal(price_raw)
                assert price > 0
            except Exception:
                messages.error(request, "To'g'ri narx kiriting.")
                return redirect('manage_products')

            if name:
                if Product.objects.filter(name=name).exists():
                    messages.warning(request, f"'{name}' allaqachon mavjud.")
                else:
                    cat = ProductCategory.objects.filter(id=cat_id).first() if cat_id else None
                    p = Product.objects.create(name=name, price=price, category=cat, image=image)
                    FinishedGoodsInventory.objects.create(product=p, stock=0)
                    messages.success(request, f"Mahsulot '{name}' qo'shildi.")
            else:
                messages.error(request, "Mahsulot nomi majburiy.")

        elif action == 'edit_product':
            pid = request.POST.get('product_id')
            try:
                p = Product.objects.get(id=pid)
                p.name = request.POST.get('name', p.name).strip()
                price_raw = request.POST.get('price', '').strip()
                if price_raw:
                    p.price = Decimal(price_raw)
                cat_id = request.POST.get('category', '')
                p.category = ProductCategory.objects.filter(id=cat_id).first() if cat_id else None
                if request.FILES.get('image'):
                    p.image = request.FILES['image']
                p.save()
                messages.success(request, f"'{p.name}' yangilandi.")
            except Product.DoesNotExist:
                messages.error(request, "Mahsulot topilmadi.")

        elif action == 'add_material':
            name = request.POST.get('name', '').strip()
            unit = request.POST.get('unit', '').strip()
            stock_raw = request.POST.get('stock', '0').strip()
            try:
                stock = Decimal(stock_raw)
            except Exception:
                stock = Decimal('0')
            if name and unit:
                mat, created = RawMaterial.objects.get_or_create(name=name, defaults={'unit': unit, 'stock': stock})
                if created:
                    messages.success(request, f"Xom ashyo '{name}' qo'shildi.")
                else:
                    messages.warning(request, f"'{name}' allaqachon mavjud.")
            else:
                messages.error(request, "Nom va birlik majburiy.")

        elif action == 'add_recipe':
            product_id = request.POST.get('recipe_product')
            ingredient_ids = request.POST.getlist('ingredient_id')
            ingredient_qtys = request.POST.getlist('ingredient_qty')
            ingredient_grams = request.POST.getlist('ingredient_qty_g')

            if product_id:
                try:
                    with transaction.atomic():
                        product = Product.objects.get(id=product_id)
                        recipe, _ = Recipe.objects.get_or_create(product=product, defaults={'batch_size': 1})
                        recipe.batch_size = 1
                        recipe.save()
                        recipe.items.all().delete()
                        added = 0
                        grams_list = ingredient_grams or [''] * len(ingredient_ids)
                        for mat_id, qty, grams in zip(ingredient_ids, ingredient_qtys, grams_list):
                            if not mat_id:
                                continue
                            grams_d = Decimal(grams) if grams.strip() else None
                            if qty.strip():
                                qty_d = Decimal(qty)
                            elif grams_d:
                                # gramm/ml kiritilgan, qty bo'sh — 0 saqlaymiz, ishlab chiqarishda grams ishlatiladi
                                qty_d = Decimal('0')
                            else:
                                continue
                            RecipeItem.objects.create(
                                recipe=recipe,
                                raw_material_id=mat_id,
                                quantity=qty_d,
                                quantity_grams=grams_d,
                            )
                            added += 1
                        if added == 0:
                            raise ValueError("Kamida bitta ingredient kerak.")
                    messages.success(request, f"'{product.name}' retsepti saqlandi (1 dona uchun).")
                except (Product.DoesNotExist, ValueError, AssertionError) as e:
                    messages.error(request, str(e) or "Xato.")
            else:
                messages.error(request, "Mahsulotni tanlang.")

        elif action == 'edit_recipe':
            recipe_id = request.POST.get('recipe_id')
            ingredient_ids = request.POST.getlist('ingredient_id')
            ingredient_qtys = request.POST.getlist('ingredient_qty')
            ingredient_grams = request.POST.getlist('ingredient_qty_g')
            try:
                recipe = Recipe.objects.get(id=recipe_id)
                recipe.batch_size = 1
                recipe.save()
                recipe.items.all().delete()
                grams_list = ingredient_grams or [''] * len(ingredient_ids)
                for mat_id, qty, grams in zip(ingredient_ids, ingredient_qtys, grams_list):
                    if not mat_id:
                        continue
                    grams_d = Decimal(grams) if grams.strip() else None
                    if qty.strip():
                        qty_d = Decimal(qty)
                    elif grams_d:
                        qty_d = Decimal('0')
                    else:
                        continue
                    RecipeItem.objects.create(
                        recipe=recipe,
                        raw_material_id=mat_id,
                        quantity=qty_d,
                        quantity_grams=grams_d,
                    )
                messages.success(request, "Retsept tahrirlandi (1 dona uchun).")
            except Recipe.DoesNotExist:
                messages.error(request, "Retsept topilmadi.")

        elif action == 'delete_product':
            pid = request.POST.get('product_id')
            try:
                Product.objects.get(id=pid).delete()
                messages.success(request, "Mahsulot o'chirildi.")
            except Product.DoesNotExist:
                messages.error(request, "Topilmadi.")

        elif action == 'edit_material':
            mid = request.POST.get('material_id')
            try:
                m = RawMaterial.objects.get(id=mid)
                new_name = request.POST.get('name', '').strip()
                new_unit = request.POST.get('unit', '').strip()
                new_stock = request.POST.get('stock', '').strip()
                if new_name:
                    m.name = new_name
                if new_unit:
                    m.unit = new_unit
                if new_stock != '':
                    m.stock = Decimal(new_stock)
                m.save()
                messages.success(request, f"'{m.name}' yangilandi.")
            except (RawMaterial.DoesNotExist, InvalidOperation) as e:
                messages.error(request, str(e))

        elif action == 'delete_material':
            mid = request.POST.get('material_id')
            try:
                RawMaterial.objects.get(id=mid).delete()
                messages.success(request, "Xom ashyo o'chirildi.")
            except RawMaterial.DoesNotExist:
                messages.error(request, "Topilmadi.")

        elif action == 'receive_material':
            mid = request.POST.get('material_id')
            qty = request.POST.get('qty', '0').strip()
            try:
                m = RawMaterial.objects.get(id=mid)
                qty_d = Decimal(qty)
                if qty_d <= 0:
                    raise ValueError("Miqdor 0 dan katta bo'lishi kerak.")
                m.stock += qty_d
                m.save()
                messages.success(request, f"{m.name}: {qty_d} {m.unit} qabul qilindi.")
            except (RawMaterial.DoesNotExist, InvalidOperation, ValueError) as e:
                messages.error(request, str(e))

        elif action == 'edit_balance_row':
            pid = request.POST.get('product_id')
            raw_date = request.POST.get('balance_date', '').strip()
            physical_raw = request.POST.get('physical', '').strip()
            opening_raw = request.POST.get('opening', '').strip()
            closing_raw = request.POST.get('closing', '').strip()
            try:
                d = date.fromisoformat(raw_date) if raw_date else timezone.localdate()
                # Fizik qoldiqni FinishedGoodsInventory da yangilash
                if physical_raw != '':
                    fg = FinishedGoodsInventory.objects.filter(product_id=pid).first()
                    if fg:
                        fg.stock = int(physical_raw)
                        fg.save()
                # ProductDayBalance ni yangilash
                row, _ = ProductDayBalance.objects.get_or_create(
                    balance_date=d, product_id=pid,
                    defaults={'opening_qty': int(opening_raw) if opening_raw else 0}
                )
                if opening_raw != '':
                    row.opening_qty = int(opening_raw)
                if closing_raw != '':
                    row.closing_qty = int(closing_raw)
                    row.closed_at = timezone.now()
                row.save()
                messages.success(request, "Qoldiq yangilandi.")
            except (ValueError, TypeError) as e:
                messages.error(request, str(e))
            return redirect(f'/production/manage/?balance_date={raw_date}#tab-day-balance')

        elif action == 'close_inventory_day':
            raw = request.POST.get('balance_date', '').strip()
            d = timezone.localdate()
            if raw:
                try:
                    d = date.fromisoformat(raw)
                except ValueError:
                    messages.error(request, "Sana noto'g'ri.")
                    return redirect('manage_products')
            _run_close_inventory_day(d)
            manage_redirect_balance_date = d.isoformat()
            messages.success(request, f"{d.strftime('%d.%m.%Y')} bo'yicha kun yakuni yozildi (joriy ombor asosida).")

        elif action == 'carry_inventory_forward':
            raw = request.POST.get('balance_date', '').strip()
            d = timezone.localdate()
            if raw:
                try:
                    d = date.fromisoformat(raw)
                except ValueError:
                    messages.error(request, "Sana noto'g'ri.")
                    return redirect('manage_products')
            next_d = d + timedelta(days=1)
            _run_carry_inventory_forward(d, next_d)
            manage_redirect_balance_date = next_d.isoformat()
            messages.success(
                request,
                f"{next_d.strftime('%d.%m.%Y')} uchun kirish qoldiqlari yangilandi (oldingi kun yopilishidan).",
            )

        if manage_redirect_balance_date:
            return redirect(f'/production/manage/?balance_date={manage_redirect_balance_date}')
        # edit_material dan keyin tab-add-material ga qaytish
        if action in ('edit_material', 'add_material', 'receive_material', 'delete_material'):
            return redirect('/production/manage/#tab-add-material')
        return redirect('manage_products')

    # ── Date filter ───────────────────────────────────────────────────
    date_from, date_to = _date_range(request)

    exp = request.GET.get('export')
    if exp == 'excel':
        return _export_recipes_excel(recipes)
    if exp == 'catalog':
        return _export_products_catalog(products)

    low_materials = raw_materials.filter(stock__lt=10)
    inventory_value = sum(
        (item.product.price * item.stock)
        for item in FinishedGoodsInventory.objects.select_related('product')
    ) or Decimal('0')

    product_count = products.count()
    recipe_count = recipes.count()
    recipe_coverage = round((recipe_count / product_count) * 100, 1) if product_count else 0

    balance_view_date = timezone.localdate()
    raw_bal = request.GET.get('balance_date', '').strip()
    if raw_bal:
        try:
            balance_view_date = date.fromisoformat(raw_bal)
        except ValueError:
            pass
    balance_prev_date = balance_view_date - timedelta(days=1)
    balance_rows = []
    for inv in FinishedGoodsInventory.objects.select_related('product').order_by('product__name'):
        b = ProductDayBalance.objects.filter(balance_date=balance_view_date, product_id=inv.product_id).first()
        balance_rows.append({
            'product': inv.product,
            'physical': inv.stock,
            'opening': b.opening_qty if b else None,
            'closing': b.closing_qty if b else None,
            'closed_at': b.closed_at if b else None,
        })

    context = {
        'products': products,
        'raw_materials': raw_materials,
        'recipes': recipes,
        'categories': categories,
        'low_materials': low_materials,
        'inventory_value': inventory_value,
        'recipe_coverage': recipe_coverage,
        'product_count': product_count,
        'material_count': raw_materials.count(),
        'recipe_count': recipe_count,
        'materials_low_count': low_materials.count(),
        'products_without_recipe': products.filter(recipe__isnull=True).count(),
        'low_material_preview': low_materials.order_by('stock')[:6],
        'avg_material_stock': raw_materials.aggregate(avg=Avg('stock'))['avg'] or 0,
        'date_from': date_from,
        'date_to': date_to,
        'period': request.GET.get('period', 'today'),
        'balance_view_date': balance_view_date,
        'balance_prev_date': balance_prev_date,
        'balance_rows': balance_rows,
    }
    return render(request, 'manage_products.html', context)


def _export_products_catalog(products):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"
    fill = PatternFill("solid", fgColor="D4A373")
    bold = Font(bold=True)
    headers = ['Mahsulot', 'Kategoriya', 'Narx (UZS)', 'Qoldiq (dona)', 'Ishlab chiqarilgan']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    qs = products.select_related('category', 'inventory')
    for row_idx, p in enumerate(qs, 2):
        inv = getattr(p, 'inventory', None)
        ws.cell(row=row_idx, column=1, value=p.name)
        ws.cell(row=row_idx, column=2, value=p.category.name if p.category else '')
        ws.cell(row=row_idx, column=3, value=float(p.price))
        ws.cell(row=row_idx, column=4, value=inv.stock if inv else 0)
        prod_at = ''
        if inv and inv.produced_at:
            prod_at = timezone.localtime(inv.produced_at).strftime('%d.%m.%Y %H:%M')
        ws.cell(row=row_idx, column=5, value=prod_at)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="products_catalog.xlsx"'
    return resp


def _export_recipes_excel(recipes):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retseptlar"
    fill = PatternFill("solid", fgColor="D4A373")
    bold = Font(bold=True)
    headers = ['Mahsulot', 'Partiya hajmi', 'Ingredient', 'Miqdor', 'Birlik', 'Gramm (ixtiyoriy)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    row_idx = 2
    for recipe in recipes:
        for item in recipe.items.all():
            ws.cell(row=row_idx, column=1, value=recipe.product.name)
            ws.cell(row=row_idx, column=2, value=recipe.batch_size)
            ws.cell(row=row_idx, column=3, value=item.raw_material.name)
            ws.cell(row=row_idx, column=4, value=float(item.quantity))
            ws.cell(row=row_idx, column=5, value=item.raw_material.unit)
            ws.cell(row=row_idx, column=6, value=float(item.quantity_grams) if item.quantity_grams else '')
            row_idx += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="recipes.xlsx"'
    return resp


@login_required
def recipe_print(request, recipe_id):
    """Print a single recipe as HTML."""
    if not _can_access(request.user, 'seller', 'production_manager'):
        return JsonResponse({'ok': False}, status=403)
    
    try:
        recipe = Recipe.objects.select_related('product').prefetch_related('items__raw_material').get(id=recipe_id)
    except Recipe.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Topilmadi'}, status=404)
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="uz">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Retsept: {recipe.product.name}</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; }}
            .print-container {{
                background: white;
                padding: 30px;
                max-width: 800px;
                margin: 0 auto;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }}
            h1 {{ color: #2f2a25; margin-bottom: 10px; font-size: 28px; }}
            .recipe-meta {{ display: flex; gap: 20px; margin-bottom: 20px; color: #666; font-size: 14px; }}
            .meta-item {{ display: flex; flex-direction: column; }}
            .meta-label {{ font-weight: bold; color: #333; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background: #d4a373; color: white; padding: 12px; text-align: left; font-weight: bold; }}
            td {{ padding: 10px 12px; border-bottom: 1px solid #ddd; }}
            tr:hover {{ background: #f9f9f9; }}
            .total-row {{ font-weight: bold; background: #efefef; }}
            @media print {{
                body {{ background: white; padding: 0; }}
                .print-container {{ box-shadow: none; }}
            }}
            .print-btn {{
                background: #d4a373;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                margin-bottom: 20px;
                font-size: 14px;
            }}
            .print-btn:hover {{ background: #b88656; }}
        </style>
    </head>
    <body>
        <div class="print-container">
            <button class="print-btn" onclick="window.print()">🖨️ Chop etish</button>
            <h1>{recipe.product.name}</h1>
            <div class="recipe-meta">
                <div class="meta-item">
                    <span class="meta-label">Partiya hajmi:</span>
                    <span>{recipe.batch_size} dona</span>
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Ingredientlar</th>
                        <th>Miqdor</th>
                        <th>Birlik</th>
                        <th>Gramm (ixtiyoriy)</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for idx, item in enumerate(recipe.items.all(), 1):
        grams_val = f"{item.quantity_grams}" if item.quantity_grams else "—"
        html_content += f"""
                    <tr>
                        <td>{idx}</td>
                        <td>{item.raw_material.name}</td>
                        <td>{item.quantity}</td>
                        <td>{item.raw_material.unit}</td>
                        <td>{grams_val}</td>
                    </tr>
        """
    
    html_content += """
                </tbody>
            </table>
        </div>
    </body>
    </html>
    """
    
    return HttpResponse(html_content, content_type='text/html')


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def _xlsx_header(ws, headers, color='D4A373'):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill('solid', fgColor=color)
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold; c.fill = fill
        c.alignment = Alignment(horizontal='center')
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# RETSEPT — JSON export / import
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def recipe_export_json(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return JsonResponse({'ok': False}, status=403)
    data = []
    for recipe in Recipe.objects.select_related('product').prefetch_related('items__raw_material'):
        data.append({
            'product': recipe.product.name,
            'batch_size': recipe.batch_size,
            'items': [
                {
                    'ingredient': item.raw_material.name,
                    'unit': item.raw_material.unit,
                    'qty': str(item.quantity),
                    'grams': str(item.quantity_grams) if item.quantity_grams else None,
                }
                for item in recipe.items.all()
            ],
        })
    resp = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
    resp['Content-Disposition'] = 'attachment; filename="recipes.json"'
    return resp


@login_required
def recipe_import_json(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('manage_products')
    f = request.FILES.get('recipe_json_file')
    if not f:
        messages.error(request, "JSON fayl tanlanmadi.")
        return redirect('/production/manage/?tab=tab-recipes')
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception as e:
        messages.error(request, f"JSON o'qishda xato: {e}")
        return redirect('/production/manage/?tab=tab-recipes')
    created = updated = 0
    with transaction.atomic():
        for entry in data:
            prod_name = str(entry.get('product', '')).strip()
            if not prod_name:
                continue
            product = Product.objects.filter(name=prod_name).first()
            if not product:
                product = Product.objects.create(name=prod_name, price=Decimal('0'))
                FinishedGoodsInventory.objects.create(product=product, stock=0)
            recipe, is_new = Recipe.objects.get_or_create(product=product, defaults={'batch_size': entry.get('batch_size', 1)})
            recipe.batch_size = entry.get('batch_size', 1)
            recipe.save()
            recipe.items.all().delete()
            for item in entry.get('items', []):
                mat, _ = RawMaterial.objects.get_or_create(
                    name=str(item['ingredient']).strip(),
                    defaults={'unit': item.get('unit', 'kg'), 'stock': Decimal('0')}
                )
                RecipeItem.objects.create(
                    recipe=recipe, raw_material=mat,
                    quantity=Decimal(str(item.get('qty', 0))),
                    quantity_grams=Decimal(str(item['grams'])) if item.get('grams') else None,
                )
            if is_new: created += 1
            else: updated += 1
    messages.success(request, f"JSON import: {created} yangi, {updated} yangilandi.")
    return redirect('/production/manage/?tab=tab-recipes')


# ─────────────────────────────────────────────────────────────────────────────
# MAHSULOTLAR — Excel + JSON export / import
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def products_export_excel(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    import openpyxl
    wb = openpyxl.Workbook()
    ws = _xlsx_header(wb.active, ['Nomi', 'Kategoriya', 'Narx (UZS)'])
    ws.title = 'Mahsulotlar'
    for i, p in enumerate(Product.objects.select_related('category').order_by('name'), 2):
        ws.cell(i, 1, p.name)
        ws.cell(i, 2, p.category.name if p.category else '')
        ws.cell(i, 3, float(p.price))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 24
    return _xlsx_response(wb, 'products.xlsx')


@login_required
def products_export_json(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    data = [
        {'name': p.name, 'category': p.category.name if p.category else '', 'price': str(p.price)}
        for p in Product.objects.select_related('category').order_by('name')
    ]
    resp = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
    resp['Content-Disposition'] = 'attachment; filename="products.json"'
    return resp


@login_required
def products_import_excel(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('manage_products')
    import openpyxl
    f = request.FILES.get('products_excel_file')
    if not f:
        messages.error(request, "Fayl tanlanmadi.")
        return redirect('/production/manage/?tab=tab-overview')
    try:
        ws = openpyxl.load_workbook(f).active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        messages.error(request, f"Fayl xato: {e}")
        return redirect('/production/manage/?tab=tab-overview')
    created = updated = 0
    with transaction.atomic():
        for row in rows:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            cat_name = str(row[1]).strip() if row[1] else ''
            try:
                price = Decimal(str(row[2])) if row[2] else Decimal('0')
            except Exception:
                price = Decimal('0')
            cat = ProductCategory.objects.filter(name=cat_name).first() if cat_name else None
            if not cat and cat_name:
                cat = ProductCategory.objects.create(name=cat_name)
            p, is_new = Product.objects.get_or_create(name=name, defaults={'price': price, 'category': cat})
            if not is_new:
                p.price = price; p.category = cat; p.save()
                updated += 1
            else:
                FinishedGoodsInventory.objects.get_or_create(product=p)
                created += 1
    messages.success(request, f"Mahsulotlar import: {created} yangi, {updated} yangilandi.")
    return redirect('/production/manage/?tab=tab-overview')


@login_required
def products_import_json(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('manage_products')
    f = request.FILES.get('products_json_file')
    if not f:
        messages.error(request, "JSON fayl tanlanmadi.")
        return redirect('/production/manage/?tab=tab-overview')
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception as e:
        messages.error(request, f"JSON xato: {e}")
        return redirect('/production/manage/?tab=tab-overview')
    created = updated = 0
    with transaction.atomic():
        for entry in data:
            name = str(entry.get('name', '')).strip()
            if not name:
                continue
            cat_name = str(entry.get('category', '')).strip()
            try:
                price = Decimal(str(entry.get('price', 0)))
            except Exception:
                price = Decimal('0')
            cat = None
            if cat_name:
                cat, _ = ProductCategory.objects.get_or_create(name=cat_name)
            p, is_new = Product.objects.get_or_create(name=name, defaults={'price': price, 'category': cat})
            if not is_new:
                p.price = price; p.category = cat; p.save()
                updated += 1
            else:
                FinishedGoodsInventory.objects.get_or_create(product=p)
                created += 1
    messages.success(request, f"JSON import: {created} yangi, {updated} yangilandi.")
    return redirect('/production/manage/?tab=tab-overview')


# ─────────────────────────────────────────────────────────────────────────────
# XOM ASHYO — Excel + JSON export / import
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def materials_export_excel(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    import openpyxl
    wb = openpyxl.Workbook()
    ws = _xlsx_header(wb.active, ['Nomi', 'Birlik', 'Qoldiq'])
    ws.title = 'Xom ashyo'
    for i, m in enumerate(RawMaterial.objects.order_by('name'), 2):
        ws.cell(i, 1, m.name)
        ws.cell(i, 2, m.unit)
        ws.cell(i, 3, float(m.stock))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    return _xlsx_response(wb, 'materials.xlsx')


@login_required
def materials_export_json(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    data = [
        {'name': m.name, 'unit': m.unit, 'stock': str(m.stock)}
        for m in RawMaterial.objects.order_by('name')
    ]
    resp = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
    resp['Content-Disposition'] = 'attachment; filename="materials.json"'
    return resp


@login_required
def materials_import_excel(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('manage_products')
    import openpyxl
    f = request.FILES.get('materials_excel_file')
    if not f:
        messages.error(request, "Fayl tanlanmadi.")
        return redirect('/production/manage/?tab=tab-add-material')
    try:
        ws = openpyxl.load_workbook(f).active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        messages.error(request, f"Fayl xato: {e}")
        return redirect('/production/manage/?tab=tab-add-material')
    created = updated = 0
    with transaction.atomic():
        for row in rows:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            unit = str(row[1]).strip() if row[1] else 'kg'
            try:
                stock = Decimal(str(row[2])) if row[2] not in (None, '') else Decimal('0')
            except Exception:
                stock = Decimal('0')
            m, is_new = RawMaterial.objects.get_or_create(name=name, defaults={'unit': unit, 'stock': stock})
            if not is_new:
                m.unit = unit; m.stock = stock; m.save()
                updated += 1
            else:
                created += 1
    messages.success(request, f"Xom ashyo import: {created} yangi, {updated} yangilandi.")
    return redirect('/production/manage/?tab=tab-add-material')


@login_required
def materials_import_json(request):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('manage_products')
    f = request.FILES.get('materials_json_file')
    if not f:
        messages.error(request, "JSON fayl tanlanmadi.")
        return redirect('/production/manage/?tab=tab-add-material')
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception as e:
        messages.error(request, f"JSON xato: {e}")
        return redirect('/production/manage/?tab=tab-add-material')
    created = updated = 0
    with transaction.atomic():
        for entry in data:
            name = str(entry.get('name', '')).strip()
            if not name:
                continue
            unit = str(entry.get('unit', 'kg')).strip()
            try:
                stock = Decimal(str(entry.get('stock', 0)))
            except Exception:
                stock = Decimal('0')
            m, is_new = RawMaterial.objects.get_or_create(name=name, defaults={'unit': unit, 'stock': stock})
            if not is_new:
                m.unit = unit; m.stock = stock; m.save()
                updated += 1
            else:
                created += 1
    messages.success(request, f"JSON import: {created} yangi, {updated} yangilandi.")
    return redirect('/production/manage/?tab=tab-add-material')


@login_required
def recipe_import_template(request):
    """Namuna Excel fayl yuklab olish."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retseptlar"
    fill = PatternFill("solid", fgColor="D4A373")
    bold = Font(bold=True)
    headers = ['Mahsulot', 'Ingredient', 'Miqdor', 'Birlik', 'Gramm']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')
    # Namuna qatorlar
    samples = [
        ['Obi Non', 'Un', 0.5, 'kg', ''],
        ['Obi Non', 'Suv', 0.3, 'liter', ''],
        ['Obi Non', 'Tuz', 0.01, 'kg', ''],
        ['Patir Non', 'Un', 0.4, 'kg', ''],
        ['Patir Non', 'Yog', 0.05, 'kg', ''],
    ]
    for i, row in enumerate(samples, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="recipe_import_template.xlsx"'
    return resp


@login_required
def recipe_import(request):
    """Excel fayldan retseptlarni import qilish."""
    if not _can_access(request.user, 'seller', 'production_manager'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('manage_products')

    try:
        import openpyxl
    except ImportError:
        messages.error(request, "openpyxl o'rnatilmagan. pip install openpyxl")
        return redirect('/production/manage/#tab-recipes')

    f = request.FILES.get('recipe_file')
    if not f:
        messages.error(request, "Fayl tanlanmadi.")
        return redirect('/production/manage/#tab-recipes')

    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        messages.error(request, f"Faylni o'qishda xato: {e}")
        return redirect('/production/manage/#tab-recipes')

    # Mahsulot nomi bo'yicha guruhlash
    from collections import defaultdict
    recipe_data = defaultdict(list)
    for row in rows:
        if not row or not row[0]:
            continue
        prod_name = str(row[0]).strip()
        ing_name  = str(row[1]).strip() if row[1] else ''
        qty_raw   = row[2]
        unit      = str(row[3]).strip() if row[3] else 'kg'
        grams_raw = row[4]
        if not ing_name:
            continue
        try:
            qty = Decimal(str(qty_raw)) if qty_raw not in (None, '') else Decimal('0')
        except Exception:
            qty = Decimal('0')
        try:
            grams = Decimal(str(grams_raw)) if grams_raw not in (None, '') else None
        except Exception:
            grams = None
        recipe_data[prod_name].append({'ing_name': ing_name, 'qty': qty, 'unit': unit, 'grams': grams})

    created_count = 0
    updated_count = 0
    errors = []

    with transaction.atomic():
        for prod_name, items in recipe_data.items():
            product = Product.objects.filter(name=prod_name).first()
            if not product:
                # Mahsulot yo'q bo'lsa avtomatik yaratamiz
                product = Product.objects.create(name=prod_name, price=Decimal('0'))
                FinishedGoodsInventory.objects.create(product=product, stock=0)

            recipe, created = Recipe.objects.get_or_create(product=product, defaults={'batch_size': 1})
            recipe.batch_size = 1
            recipe.save()
            recipe.items.all().delete()

            added = 0
            for item in items:
                mat, _ = RawMaterial.objects.get_or_create(
                    name=item['ing_name'],
                    defaults={'unit': item['unit'], 'stock': Decimal('0')}
                )
                RecipeItem.objects.create(
                    recipe=recipe,
                    raw_material=mat,
                    quantity=item['qty'],
                    quantity_grams=item['grams'],
                )
                added += 1

            if created:
                created_count += 1
            else:
                updated_count += 1

    msg = f"Import tugadi: {created_count} ta yangi, {updated_count} ta yangilangan retsept."
    if errors:
        msg += " Xatolar: " + "; ".join(errors)
    messages.success(request, msg)
    return redirect('/production/manage/#tab-recipes')


@login_required
def recipe_json(request, recipe_id):
    if not _can_access(request.user, 'seller', 'production_manager'):
        return JsonResponse({'ok': False}, status=403)
    
    try:
        recipe = Recipe.objects.select_related('product').prefetch_related('items__raw_material').get(id=recipe_id)
        items = [
            {
                'id': item.id,
                'material_id': item.raw_material.id,
                'material_name': item.raw_material.name,
                'quantity': str(item.quantity),
                'quantity_grams': str(item.quantity_grams) if item.quantity_grams else '',
                'unit': item.raw_material.unit
            }
            for item in recipe.items.all()
        ]
        return JsonResponse({
            'ok': True,
            'recipe': {
                'id': recipe.id,
                'product_id': recipe.product.id,
                'product_name': recipe.product.name,
                'batch_size': recipe.batch_size,
            },
            'items': items
        })
    except Recipe.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Topilmadi'}, status=404)
