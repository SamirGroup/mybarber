import io
import json
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from urllib.parse import urlencode

from .models import AdvancePayment, DailyReport, Employee, Position, Shift


def _can_access(user, *roles):
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=roles).exists()


def _parse_decimal(val, default=None):
    val = (val or '').strip()
    if not val:
        return default
    try:
        return Decimal(val.replace(',', '.'))
    except InvalidOperation:
        return default


def _date_range(request):
    period = request.GET.get('period', 'today')
    today = timezone.localdate()
    if period == 'week':
        return today - timedelta(days=6), today
    if period == 'custom':
        try:
            df = date.fromisoformat(request.GET.get('date_from', ''))
            dt = date.fromisoformat(request.GET.get('date_to', ''))
            return df, dt
        except ValueError:
            pass
    return today, today


@login_required
def hr_dashboard(request):
    if not _can_access(request.user, 'hr'):
        return redirect('dashboard')

    employees = Employee.objects.select_related('shift').all()
    shifts = Shift.objects.all().order_by('name')
    positions = Position.objects.all().order_by('name')
    advances = AdvancePayment.objects.select_related('employee').order_by('-date')[:15]

    date_from, date_to = _date_range(request)

    # ── Sort ──────────────────────────────────────────────────────────
    sort_by = request.GET.get('sort', 'name')
    sort_map = {
        'name': 'name',
        'position': 'position',
        'shift': 'shift__name',
        'date_joined': 'date_joined',
        'target': '-daily_target',
    }
    employees = employees.order_by(sort_map.get(sort_by, 'name'))

    # Filter by shift
    shift_filter = request.GET.get('shift_filter', '')
    if shift_filter:
        employees = employees.filter(shift_id=shift_filter)

    # Filter by position
    pos_filter = request.GET.get('pos_filter', '')
    if pos_filter:
        employees = employees.filter(position__icontains=pos_filter)

    if request.GET.get('export') == 'excel':
        return _export_hr_excel(employees, date_from, date_to)

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'add_shift':
            name = request.POST.get('shift_name', '').strip()
            start = request.POST.get('shift_start', '')
            end = request.POST.get('shift_end', '')
            if name and start and end:
                Shift.objects.get_or_create(name=name, defaults={'start_time': start, 'end_time': end})
                messages.success(request, f"Smena '{name}' qo'shildi.")
            else:
                messages.error(request, "Smena nomi, boshlanish va tugash vaqti majburiy.")

        elif action == 'delete_shift':
            sid = request.POST.get('shift_id')
            try:
                Shift.objects.get(id=sid).delete()
                messages.success(request, "Smena o'chirildi.")
            except Shift.DoesNotExist:
                messages.error(request, "Smena topilmadi.")

        elif action == 'add_position':
            name = request.POST.get('position_name', '').strip()
            desc = request.POST.get('position_desc', '').strip()
            if name:
                _, created = Position.objects.get_or_create(name=name, defaults={'description': desc})
                if created:
                    messages.success(request, f"Lavozim '{name}' qo'shildi.")
                else:
                    messages.warning(request, f"'{name}' lavozimi allaqachon mavjud.")
            else:
                messages.error(request, "Lavozim nomi majburiy.")

        elif action == 'edit_position':
            pid = request.POST.get('position_id')
            try:
                pos = Position.objects.get(id=pid)
                new_name = request.POST.get('position_name', '').strip()
                if new_name:
                    pos.name = new_name
                pos.description = request.POST.get('position_desc', '').strip()
                pos.save()
                messages.success(request, f"Lavozim '{pos.name}' yangilandi.")
            except Position.DoesNotExist:
                messages.error(request, "Lavozim topilmadi.")

        elif action == 'delete_position':
            pid = request.POST.get('position_id')
            try:
                Position.objects.get(id=pid).delete()
                messages.success(request, "Lavozim o'chirildi.")
            except Position.DoesNotExist:
                messages.error(request, "Lavozim topilmadi.")

        elif action == 'add_employee':
            name = request.POST.get('name', '').strip()
            phone = request.POST.get('phone', '').strip()
            position = request.POST.get('position', '').strip()
            date_joined = request.POST.get('date_joined', '')
            shift_id = request.POST.get('shift_id', '')
            is_piecework = request.POST.get('is_piecework') == 'on'
            base_salary = request.POST.get('base_salary', 0) or 0
            piecework_rate = request.POST.get('piecework_rate', 0) or 0
            daily_target = request.POST.get('daily_target', 0) or 0
            # photo = request.FILES.get('photo')

            if name and position and date_joined:
                shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
                emp, created = Employee.objects.get_or_create(
                    name=name, phone=phone,
                    defaults={
                        'position': position,
                        'date_joined': date_joined,
                        'is_piecework': is_piecework,
                        'base_salary': base_salary,
                        'piecework_rate': piecework_rate,
                        'daily_target': daily_target,
                        'status': 'active',
                        'shift': shift,
                        # 'photo': photo,
                    }
                )
                if created:
                    uid = request.POST.get('user_account_id', '').strip()
                    if uid:
                        try:
                            oid = int(uid)
                            if Employee.objects.filter(user_account_id=oid).exists():
                                messages.warning(request, "Xodim qo'shildi; tanlangan login boshqa kartada — biriktirish bekor qilindi.")
                            else:
                                emp.user_account_id = oid
                                emp.save(update_fields=['user_account_id'])
                        except ValueError:
                            pass
                    messages.success(request, f"Xodim '{name}' qo'shildi.")
                else:
                    messages.warning(request, "Bu xodim allaqachon mavjud.")
            else:
                messages.error(request, "Ism, lavozim va sana majburiy.")

        elif action == 'edit_employee':
            eid = request.POST.get('employee_id')
            try:
                emp = Employee.objects.get(id=eid)
                emp.name = request.POST.get('name', emp.name).strip()
                emp.phone = request.POST.get('phone', emp.phone).strip()
                emp.position = request.POST.get('position', emp.position).strip()
                emp.date_joined = request.POST.get('date_joined', emp.date_joined)
                shift_id = request.POST.get('shift_id', '')
                emp.shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
                emp.is_piecework = request.POST.get('is_piecework') == 'on'
                emp.base_salary = request.POST.get('base_salary', emp.base_salary) or emp.base_salary
                emp.piecework_rate = request.POST.get('piecework_rate', emp.piecework_rate) or emp.piecework_rate
                emp.daily_target = request.POST.get('daily_target', emp.daily_target) or emp.daily_target
                new_status = request.POST.get('status', '').strip()
                if new_status in dict(Employee.STATUS_CHOICES):
                    emp.status = new_status
                uid = request.POST.get('user_account_id', '').strip()
                if uid:
                    try:
                        oid = int(uid)
                        if Employee.objects.filter(user_account_id=oid).exclude(pk=emp.pk).exists():
                            messages.error(request, "Bu login boshqa xodimga biriktirilgan — user_account saqlanmadi.")
                        else:
                            emp.user_account_id = oid
                    except ValueError:
                        messages.error(request, "Noto'g'ri foydalanuvchi tanlandi.")
                else:
                    emp.user_account_id = None
                # if request.FILES.get('photo'):
                #     emp.photo = request.FILES['photo']
                emp.save()
                messages.success(request, f"{emp.name} yangilandi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        elif action == 'change_status':
            eid = request.POST.get('employee_id')
            new_status = request.POST.get('new_status')
            try:
                emp = Employee.objects.get(id=eid)
                emp.status = new_status
                emp.save()
                messages.success(request, f"{emp.name} holati o'zgartirildi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        elif action == 'delete_employee':
            eid = request.POST.get('employee_id')
            try:
                Employee.objects.get(id=eid).delete()
                messages.success(request, "Xodim o'chirildi.")
            except Employee.DoesNotExist:
                messages.error(request, "Topilmadi.")

        elif action == 'add_daily_report':
            eid = request.POST.get('employee_id')
            report_date = request.POST.get('report_date', str(timezone.localdate()))
            shift_id = request.POST.get('shift_id', '')
            check_in = request.POST.get('check_in') or None
            check_out = request.POST.get('check_out') or None
            units = int(request.POST.get('units_produced', 0) or 0)
            notes = request.POST.get('notes', '')
            hours_expected = _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8')
            hours_present = _parse_decimal(request.POST.get('hours_present'))
            hours_absent = _parse_decimal(request.POST.get('hours_absent'))
            hours_left_early = _parse_decimal(request.POST.get('hours_left_early'))
            was_present = request.POST.get('attendance', 'present') == 'present'
            try:
                emp = Employee.objects.get(id=eid)
                shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
                DailyReport.objects.update_or_create(
                    employee=emp, date=report_date, shift=shift,
                    defaults={
                        'check_in': check_in,
                        'check_out': check_out,
                        'units_produced': units,
                        'notes': notes,
                        'hours_expected': hours_expected,
                        'hours_present': hours_present,
                        'hours_absent': hours_absent,
                        'hours_left_early': hours_left_early,
                        'was_present': was_present,
                    },
                )
                messages.success(request, f"{emp.name} uchun hisobot saqlandi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        elif 'give_advance' in request.POST:
            emp_id = request.POST.get('employee')
            amount = request.POST.get('amount')
            try:
                emp = Employee.objects.get(id=emp_id)
                AdvancePayment.objects.create(employee=emp, amount=amount)
                messages.success(request, f"{emp.name} ga {amount} UZS avans berildi.")
            except Employee.DoesNotExist:
                messages.error(request, "Xodim topilmadi.")

        params = {
            'period': request.GET.get('period', 'today'),
            'sort': sort_by,
            'shift_filter': shift_filter,
            'pos_filter': pos_filter,
            'date_from': request.GET.get('date_from', ''),
            'date_to': request.GET.get('date_to', ''),
        }
        return redirect('/hr/?' + urlencode(params))

    # Daily reports for date range
    daily_reports = (
        DailyReport.objects
        .select_related('employee', 'shift')
        .filter(date__gte=date_from, date__lte=date_to)
        .order_by('-date', 'employee__name')
    )

    seller_users = User.objects.filter(is_active=True).order_by('username')

    employee_edit_payload = [
        {
            'id': e.id,
            'name': e.name,
            'phone': e.phone or '',
            'position': e.position,
            'date_joined': str(e.date_joined),
            'shift_id': e.shift_id,
            'status': e.status,
            'is_piecework': bool(e.is_piecework),
            'base_salary': str(e.base_salary),
            'piecework_rate': str(e.piecework_rate),
            'daily_target': e.daily_target,
            'user_account_id': e.user_account_id,
        }
        for e in Employee.objects.select_related('shift').order_by('name')
    ]

    context = {
        'employees': employees,
        'active_employees': Employee.objects.filter(status='active').order_by('name'),
        'seller_users': seller_users,
        'employee_edit_payload': employee_edit_payload,
        'shifts': shifts,
        'positions': positions,
        'advances': advances,
        'daily_reports': daily_reports,
        'today': timezone.localdate(),
        'date_from': date_from,
        'date_to': date_to,
        'period': request.GET.get('period', 'today'),
        'sort_by': sort_by,
        'shift_filter': shift_filter,
        'pos_filter': pos_filter,
    }
    return render(request, 'hr.html', context)


@login_required
def employee_report(request, emp_id):
    if not _can_access(request.user, 'hr'):
        return redirect('dashboard')

    emp = Employee.objects.select_related('shift').get(id=emp_id)

    # --- Sana oralig'i ---
    period = request.GET.get('period', 'month')
    today = timezone.localdate()
    if period == 'today':
        date_from = date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'custom':
        try:
            date_from = date.fromisoformat(request.GET.get('date_from', ''))
            date_to = date.fromisoformat(request.GET.get('date_to', ''))
        except ValueError:
            date_from = today.replace(day=1)
            date_to = today
    else:  # month
        date_from = today.replace(day=1)
        date_to = today

    # --- POST: hisobot tahrirlash ---
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'edit_report':
            rid = request.POST.get('report_id')
            try:
                r = DailyReport.objects.get(id=rid, employee=emp)
                r.check_in = request.POST.get('check_in') or None
                r.check_out = request.POST.get('check_out') or None
                r.was_present = request.POST.get('was_present') == '1'
                r.absence_reason = request.POST.get('absence_reason', '') if not r.was_present else ''
                r.units_produced = int(request.POST.get('units_produced', 0) or 0)
                r.hours_expected = _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8')
                r.hours_present = _parse_decimal(request.POST.get('hours_present'))
                r.notes = request.POST.get('notes', '')
                r.save()
                messages.success(request, f"{r.date} hisoboti yangilandi.")
            except DailyReport.DoesNotExist:
                messages.error(request, "Hisobot topilmadi.")
        elif action == 'add_report':
            rdate = request.POST.get('report_date', str(today))
            shift_id = request.POST.get('shift_id', '')
            shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
            was_present = request.POST.get('was_present') == '1'
            absence_reason = request.POST.get('absence_reason', '') if not was_present else ''
            DailyReport.objects.update_or_create(
                employee=emp, date=rdate, shift=shift,
                defaults={
                    'check_in': request.POST.get('check_in') or None,
                    'check_out': request.POST.get('check_out') or None,
                    'was_present': was_present,
                    'absence_reason': absence_reason,
                    'units_produced': int(request.POST.get('units_produced', 0) or 0),
                    'hours_expected': _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8'),
                    'hours_present': _parse_decimal(request.POST.get('hours_present')),
                    'notes': request.POST.get('notes', ''),
                }
            )
            messages.success(request, "Hisobot saqlandi.")
        from urllib.parse import urlencode as _ue
        return redirect('/hr/employee/{}/report/?'.format(emp_id) + _ue({
            'period': period,
            'date_from': str(date_from),
            'date_to': str(date_to),
        }))

    # --- Hisobotlar ---
    reports = (
        DailyReport.objects
        .filter(employee=emp, date__gte=date_from, date__lte=date_to)
        .order_by('date')
    )

    # --- Statistika ---
    total_days = (date_to - date_from).days + 1
    present_days = reports.filter(was_present=True).count()
    absent_days = reports.filter(was_present=False).count()
    absent_sick = reports.filter(was_present=False, absence_reason='sick').count()
    absent_personal = reports.filter(was_present=False, absence_reason='personal').count()
    absent_approved = reports.filter(was_present=False, absence_reason='approved').count()
    absent_no_reason = reports.filter(was_present=False, absence_reason='no_reason').count()

    total_units = sum(
        (r.units_produced or 0) + (r.units_from_sales or 0) for r in reports
    )
    total_hours = sum(
        float(r.hours_present or 0) for r in reports
    )

    # --- Oylik hisob ---
    if emp.is_piecework:
        # Ishbay: jami dona × stavka
        salary_calc = Decimal(str(total_units)) * (emp.piecework_rate or Decimal('0'))
        salary_type = 'piecework'
    else:
        # Kunlik: kelgan kunlar × (oylik / ish kunlari)
        work_days_in_period = total_days
        if work_days_in_period > 0:
            daily_rate = (emp.base_salary or Decimal('0')) / Decimal(str(work_days_in_period))
            salary_calc = daily_rate * Decimal(str(present_days))
        else:
            salary_calc = Decimal('0')
        salary_type = 'daily'

    # Avans chegirib tashlash
    advances = AdvancePayment.objects.filter(
        employee=emp,
        date__gte=date_from,
        date__lte=date_to
    )
    total_advance = sum(a.amount for a in advances) or Decimal('0')
    net_salary = salary_calc - total_advance

    context = {
        'emp': emp,
        'reports': reports,
        'shifts': Shift.objects.all().order_by('name'),
        'date_from': date_from,
        'date_to': date_to,
        'period': period,
        'today': today,
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'absent_sick': absent_sick,
        'absent_personal': absent_personal,
        'absent_approved': absent_approved,
        'absent_no_reason': absent_no_reason,
        'total_units': total_units,
        'total_hours': round(total_hours, 2),
        'salary_calc': salary_calc,
        'salary_type': salary_type,
        'total_advance': total_advance,
        'net_salary': net_salary,
        'advances': advances,
        'absence_reasons': DailyReport.ABSENCE_REASON_CHOICES,
    }
    return render(request, 'hr_employee_report.html', context)


@login_required
def employee_report(request, emp_id):
    if not _can_access(request.user, 'hr'):
        return redirect('dashboard')
    try:
        emp = Employee.objects.select_related('shift').get(id=emp_id)
    except Employee.DoesNotExist:
        return redirect('hr_dashboard')

    today = timezone.localdate()
    period = request.GET.get('period', 'month')

    if period == 'today':
        date_from = date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'custom':
        try:
            date_from = date.fromisoformat(request.GET.get('date_from', ''))
            date_to = date.fromisoformat(request.GET.get('date_to', ''))
        except ValueError:
            date_from = today.replace(day=1)
            date_to = today
    else:  # month
        date_from = today.replace(day=1)
        date_to = today

    # POST: hisobot qo'shish yoki tahrirlash
    if request.method == 'POST':
        action = request.POST.get('action', '')
        shift_id = request.POST.get('shift_id', '')
        shift = Shift.objects.filter(id=shift_id).first() if shift_id else None
        was_present = request.POST.get('was_present') == '1'
        absence_reason = '' if was_present else request.POST.get('absence_reason', '')
        defaults = {
            'check_in': request.POST.get('check_in') or None,
            'check_out': request.POST.get('check_out') or None,
            'was_present': was_present,
            'absence_reason': absence_reason,
            'units_produced': int(request.POST.get('units_produced', 0) or 0),
            'hours_expected': _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8'),
            'hours_present': _parse_decimal(request.POST.get('hours_present')),
            'notes': request.POST.get('notes', ''),
        }
        if action == 'edit_report':
            rid = request.POST.get('report_id')
            try:
                r = DailyReport.objects.get(id=rid, employee=emp)
                for k, v in defaults.items():
                    setattr(r, k, v)
                r.save()
                messages.success(request, f"{r.date} hisoboti yangilandi.")
            except DailyReport.DoesNotExist:
                messages.error(request, "Hisobot topilmadi.")
        elif action == 'add_report':
            rdate = request.POST.get('report_date', str(today))
            DailyReport.objects.update_or_create(
                employee=emp, date=rdate, shift=shift,
                defaults=defaults
            )
            messages.success(request, "Hisobot saqlandi.")
        return redirect(f'/hr/employee/{emp_id}/report/?period={period}&date_from={date_from}&date_to={date_to}')

    reports = (
        DailyReport.objects
        .filter(employee=emp, date__gte=date_from, date__lte=date_to)
        .order_by('date')
    )

    present_days = reports.filter(was_present=True).count()
    absent_days  = reports.filter(was_present=False).count()
    absent_sick      = reports.filter(was_present=False, absence_reason='sick').count()
    absent_personal  = reports.filter(was_present=False, absence_reason='personal').count()
    absent_approved  = reports.filter(was_present=False, absence_reason='approved').count()
    absent_no_reason = reports.filter(was_present=False, absence_reason='no_reason').count()

    total_units = sum((r.units_produced or 0) + (r.units_from_sales or 0) for r in reports)
    total_hours = round(sum(float(r.hours_present or 0) for r in reports), 2)

    # Oylik hisob
    total_days = (date_to - date_from).days + 1
    daily_rate = Decimal('0')
    if emp.is_piecework:
        salary_calc = Decimal(str(total_units)) * (emp.piecework_rate or Decimal('0'))
        salary_type = 'piecework'
    else:
        daily_rate = (emp.base_salary or Decimal('0')) / Decimal(str(total_days)) if total_days else Decimal('0')
        salary_calc = daily_rate * Decimal(str(present_days))
        salary_type = 'daily'

    advances = AdvancePayment.objects.filter(
        employee=emp,
        date__gte=date_from,
        date__lte=date_to
    )
    total_advance = sum(a.amount for a in advances) or Decimal('0')
    net_salary = salary_calc - total_advance

    context = {
        'emp': emp,
        'reports': reports,
        'shifts': Shift.objects.all().order_by('name'),
        'date_from': date_from,
        'date_to': date_to,
        'period': period,
        'today': today,
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'absent_sick': absent_sick,
        'absent_personal': absent_personal,
        'absent_approved': absent_approved,
        'absent_no_reason': absent_no_reason,
        'total_units': total_units,
        'total_hours': round(total_hours, 2),
        'salary_calc': salary_calc,
        'salary_type': salary_type,
        'daily_rate': daily_rate,
        'total_advance': total_advance,
        'net_salary': net_salary,
        'advances': advances,
        'absence_reasons': DailyReport.ABSENCE_REASON_CHOICES,
    }
    return render(request, 'hr_employee_report.html', context)


@login_required
def employee_report(request, emp_id):
    if not _can_access(request.user, 'hr'):
        return redirect('dashboard')
    try:
        emp = Employee.objects.select_related('shift').get(id=emp_id)
    except Employee.DoesNotExist:
        return redirect('hr_dashboard')

    today = timezone.localdate()
    period = request.GET.get('period', 'month')

    if period == 'today':
        date_from = date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=6)
        date_to = today
    elif period == 'custom':
        try:
            date_from = date.fromisoformat(request.GET.get('date_from', ''))
            date_to   = date.fromisoformat(request.GET.get('date_to', ''))
        except ValueError:
            date_from = today.replace(day=1)
            date_to   = today
    else:  # month
        date_from = today.replace(day=1)
        date_to   = today

    if request.method == 'POST':
        action   = request.POST.get('action', '')
        shift_id = request.POST.get('shift_id', '')
        shift    = Shift.objects.filter(id=shift_id).first() if shift_id else None
        was_present    = request.POST.get('was_present') == '1'
        absence_reason = '' if was_present else request.POST.get('absence_reason', '')
        defaults = {
            'check_in':       request.POST.get('check_in') or None,
            'check_out':      request.POST.get('check_out') or None,
            'was_present':    was_present,
            'absence_reason': absence_reason,
            'units_produced': int(request.POST.get('units_produced', 0) or 0),
            'hours_expected': _parse_decimal(request.POST.get('hours_expected'), Decimal('8')) or Decimal('8'),
            'hours_present':  _parse_decimal(request.POST.get('hours_present')),
            'notes':          request.POST.get('notes', ''),
        }
        if action == 'edit_report':
            rid = request.POST.get('report_id')
            try:
                r = DailyReport.objects.get(id=rid, employee=emp)
                for k, v in defaults.items():
                    setattr(r, k, v)
                r.save()
                messages.success(request, f"{r.date} hisoboti yangilandi.")
            except DailyReport.DoesNotExist:
                messages.error(request, "Hisobot topilmadi.")
        elif action == 'add_report':
            rdate = request.POST.get('report_date', str(today))
            DailyReport.objects.update_or_create(
                employee=emp, date=rdate, shift=shift,
                defaults=defaults
            )
            messages.success(request, "Hisobot saqlandi.")
        return redirect(f'/hr/employee/{emp_id}/report/?period={period}&date_from={date_from}&date_to={date_to}')

    reports = (
        DailyReport.objects
        .filter(employee=emp, date__gte=date_from, date__lte=date_to)
        .order_by('date')
    )

    present_days     = reports.filter(was_present=True).count()
    absent_days      = reports.filter(was_present=False).count()
    absent_sick      = reports.filter(was_present=False, absence_reason='sick').count()
    absent_personal  = reports.filter(was_present=False, absence_reason='personal').count()
    absent_approved  = reports.filter(was_present=False, absence_reason='approved').count()
    absent_no_reason = reports.filter(was_present=False, absence_reason='no_reason').count()

    total_units = sum((r.units_produced or 0) + (r.units_from_sales or 0) for r in reports)
    total_hours = round(sum(float(r.hours_present or 0) for r in reports), 2)
    total_days  = (date_to - date_from).days + 1

    if emp.is_piecework:
        salary_calc = Decimal(str(total_units)) * (emp.piecework_rate or Decimal('0'))
        salary_type = 'piecework'
    else:
        daily_rate  = (emp.base_salary or Decimal('0')) / Decimal(str(total_days)) if total_days else Decimal('0')
        salary_calc = daily_rate * Decimal(str(present_days))
        salary_type = 'daily'

    advances      = AdvancePayment.objects.filter(employee=emp, date__gte=date_from, date__lte=date_to)
    total_advance = sum(a.amount for a in advances) or Decimal('0')
    net_salary    = salary_calc - total_advance

    context = {
        'emp': emp,
        'reports': reports,
        'shifts': Shift.objects.all().order_by('name'),
        'date_from': date_from,
        'date_to': date_to,
        'period': period,
        'today': today,
        'total_days': total_days,
        'present_days': present_days,
        'absent_days': absent_days,
        'absent_sick': absent_sick,
        'absent_personal': absent_personal,
        'absent_approved': absent_approved,
        'absent_no_reason': absent_no_reason,
        'total_units': total_units,
        'total_hours': total_hours,
        'salary_calc': salary_calc,
        'salary_type': salary_type,
        'total_advance': total_advance,
        'net_salary': net_salary,
        'advances': advances,
        'absence_reasons': DailyReport.ABSENCE_REASON_CHOICES,
    }
    return render(request, 'hr_employee_report.html', context)


@login_required
def positions_export_json(request):
    if not _can_access(request.user, 'hr'):
        return redirect('dashboard')
    data = [
        {'name': p.name, 'description': p.description}
        for p in Position.objects.order_by('name')
    ]
    resp = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json'
    )
    resp['Content-Disposition'] = 'attachment; filename="positions.json"'
    return resp


@login_required
def positions_import_json(request):
    if not _can_access(request.user, 'hr'):
        return redirect('dashboard')
    if request.method != 'POST':
        return redirect('hr_dashboard')
    f = request.FILES.get('positions_json_file')
    if not f:
        messages.error(request, "JSON fayl tanlanmadi.")
        return redirect('/hr/?tab=positions')
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception as e:
        messages.error(request, f"JSON o'qishda xato: {e}")
        return redirect('/hr/?tab=positions')
    created = updated = 0
    for entry in data:
        name = str(entry.get('name', '')).strip()
        if not name:
            continue
        desc = str(entry.get('description', '')).strip()
        pos, is_new = Position.objects.get_or_create(name=name, defaults={'description': desc})
        if not is_new:
            pos.description = desc
            pos.save(update_fields=['description'])
            updated += 1
        else:
            created += 1
    messages.success(request, f"Import tugadi: {created} yangi, {updated} yangilandi.")
    return redirect('/hr/?tab=positions')


def _export_hr_excel(employees, date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl o'rnatilmagan.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xodimlar"
    fill = PatternFill("solid", fgColor="D4A373")
    bold = Font(bold=True)
    headers = ['Ism', 'Lavozim', 'Smena', 'Telefon', 'Qo\'shilgan sana', 'Holat', 'Nagruzka (kun)', 'Ishbay stavka']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

    for row_idx, emp in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp.name)
        ws.cell(row=row_idx, column=2, value=emp.position)
        ws.cell(row=row_idx, column=3, value=emp.shift.name if emp.shift else '')
        ws.cell(row=row_idx, column=4, value=emp.phone)
        ws.cell(row=row_idx, column=5, value=str(emp.date_joined))
        ws.cell(row=row_idx, column=6, value=emp.get_status_display())
        ws.cell(row=row_idx, column=7, value=emp.daily_target)
        ws.cell(row=row_idx, column=8, value=float(emp.piecework_rate))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="employees_{date_from}_{date_to}.xlsx"'
    return resp
